import datetime
import openpyxl
from .models import secs_to_hhmmss


# ─── Primitive parsers ────────────────────────────────────────────────────────

def parse_timedelta_to_seconds(value):
    if value is None or value == '-' or value == '':
        return None
    if isinstance(value, datetime.timedelta):
        return value.total_seconds()
    if isinstance(value, datetime.time):
        return value.hour * 3600 + value.minute * 60 + value.second
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        if value == '-':
            return None
        parts = value.split(':')
        if len(parts) == 3:
            try:
                h, m, s = int(parts[0]), int(parts[1]), float(parts[2])
                return h * 3600 + m * 60 + s
            except ValueError:
                return None
    return None


def parse_float(value):
    if value is None or value == '-' or value == '':
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def parse_date_value(value):
    """
    Parse a date value from an Excel cell.
    openpyxl returns datetime.date / datetime.datetime for real date cells.
    String cells may contain 'DD.MM.YYYY', 'DD.MM.YY', 'DD.MM', etc.
    Returns (date_str 'DD.MM', record_date datetime.date | None, year int | None).
    """
    if value is None:
        return '', None, None

    if isinstance(value, datetime.datetime):
        d = value.date()
        return d.strftime('%d.%m'), d, d.year
    if isinstance(value, datetime.date):
        return value.strftime('%d.%m'), value, value.year

    s = str(value).strip()
    if not s or s.lower() in ('none', '-', ''):
        return '', None, None

    for fmt in ('%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d', '%d.%m.%y', '%d-%m-%Y'):
        try:
            d = datetime.datetime.strptime(s, fmt).date()
            return d.strftime('%d.%m'), d, d.year
        except ValueError:
            pass

    # Possibly day.month only
    for sep in ('.', '/'):
        parts = s.split(sep)
        if len(parts) == 2:
            try:
                day, month = int(parts[0]), int(parts[1])
                return f'{day:02d}.{month:02d}', None, None
            except ValueError:
                pass

    return s, None, None


# ─── Anomaly detection ────────────────────────────────────────────────────────

def detect_anomalies(record):
    anomalies = []

    engine_time_sec    = record.get('engine_time_sec', 0) or 0
    engine_no_move_sec = record.get('engine_no_move_sec', 0) or 0
    engine_idle_sec    = record.get('engine_idle_sec', 0) or 0
    fuel_actual        = record.get('fuel_actual')
    fuel_norm          = record.get('fuel_norm', 0) or 0
    mileage            = record.get('mileage')
    engine_time_hours  = engine_time_sec / 3600

    if fuel_actual is not None and fuel_actual < 0:
        anomalies.append(f'Отрицательный расход топлива: {fuel_actual} л')

    if engine_time_hours >= 8 and fuel_actual is not None and fuel_actual == 0:
        anomalies.append(
            f'Техника работала {engine_time_hours:.1f} ч при нулевом расходе топлива'
        )
    elif engine_time_sec > 600 and fuel_actual is not None and fuel_actual == 0:
        anomalies.append(
            f'Работа двигателя {engine_time_hours:.1f} ч без расхода топлива'
        )

    if engine_time_sec > 0 and engine_idle_sec > engine_time_sec:
        anomalies.append(
            f'Холостой ход ({secs_to_hhmmss(engine_idle_sec)}) '
            f'превышает время работы двигателя ({secs_to_hhmmss(engine_time_sec)})'
        )

    if engine_time_sec > 0 and engine_no_move_sec > engine_time_sec:
        anomalies.append(
            f'Время без движения ({secs_to_hhmmss(engine_no_move_sec)}) '
            f'превышает время работы двигателя ({secs_to_hhmmss(engine_time_sec)})'
        )

    if fuel_norm > 0 and fuel_actual is not None and fuel_actual > 0 and engine_time_hours > 0:
        actual_rate = fuel_actual / engine_time_hours
        if actual_rate > fuel_norm * 3:
            anomalies.append(
                f'Расход топлива ({actual_rate:.1f} л/ч) превышает норму '
                f'более чем в 3 раза (норма: {fuel_norm} л/ч)'
            )

    if engine_time_sec == 0 and fuel_actual is not None and fuel_actual > 0:
        anomalies.append(
            f'Зафиксирован расход топлива ({fuel_actual} л) '
            f'при нулевом времени работы двигателя'
        )

    if mileage is not None:
        if mileage > 10 and fuel_actual is not None and fuel_actual == 0:
            anomalies.append(f'Пробег {mileage:.1f} км при нулевом расходе топлива')
        if mileage > 10 and engine_time_sec == 0:
            anomalies.append(
                f'Пробег {mileage:.1f} км при нулевом времени работы двигателя'
            )
        if mileage == 0 and engine_time_sec > 3600 and fuel_actual is not None and fuel_actual > 0:
            anomalies.append(
                f'Нулевой пробег при работе двигателя {engine_time_hours:.1f} ч '
                f'и расходе {fuel_actual} л'
            )

    return len(anomalies) > 0, anomalies


# ─── Metrics ──────────────────────────────────────────────────────────────────

def calculate_metrics(record_data, report):
    """
    1. fuel_efficiency  = fuel_actual / (fuel_norm * daily_norm_hours)
    2. equipment_output = engine_time_sec / daily_norm_sec
    3. type_efficiency  (group-specific)
    """
    engine_time_sec    = record_data.get('engine_time_sec', 0) or 0
    engine_no_move_sec = record_data.get('engine_no_move_sec', 0) or 0
    engine_idle_sec    = record_data.get('engine_idle_sec', 0) or 0
    fuel_actual        = record_data.get('fuel_actual')
    fuel_norm          = record_data.get('fuel_norm', 0) or 0
    downtime_sec       = record_data.get('downtime_sec')
    group              = record_data.get('group', '')

    daily_norm_hours = report.daily_norm_sec / 3600 if report.daily_norm_sec > 0 else 0

    fuel_efficiency = None
    if (fuel_actual is not None and fuel_actual >= 0
            and daily_norm_hours > 0 and fuel_norm > 0):
        denominator = fuel_norm * daily_norm_hours
        if denominator > 0:
            fuel_efficiency = fuel_actual / denominator

    equipment_output = None
    if report.daily_norm_sec > 0:
        equipment_output = engine_time_sec / report.daily_norm_sec

    type_efficiency = None
    if group in ('Бульдозеры', 'Погрузчики'):
        if report.bulldozer_norm_sec > 0:
            type_efficiency = engine_idle_sec / report.bulldozer_norm_sec
    elif group == 'Экскаваторы':
        if downtime_sec is not None and report.excavator_norm_sec > 0:
            type_efficiency = downtime_sec / report.excavator_norm_sec
    elif group == 'Самосвалы':
        if report.dumptruck_norm_sec > 0:
            type_efficiency = engine_no_move_sec / report.dumptruck_norm_sec

    return {
        'fuel_efficiency':  fuel_efficiency,
        'equipment_output': equipment_output,
        'type_efficiency':  type_efficiency,
    }


# ─── Column detection ─────────────────────────────────────────────────────────

def _detect_header_row(ws, max_scan_row=30):
    """
    Scan worksheet rows to find the header row.
    Identifies it by presence of known column-name keywords.
    Returns the best-matching row (tuple of cell values) or None.
    """
    SIGNALS = {
        'транспортное', 'средство', 'двигател', 'расход',
        'группа', 'дата', 'смена', 'пробег',
    }
    best_row   = None
    best_score = 0

    for row in ws.iter_rows(min_row=1, max_row=max_scan_row, values_only=True):
        row_text = ' '.join(str(c).lower() for c in row if c is not None)
        score = sum(1 for kw in SIGNALS if kw in row_text)
        if score > best_score:
            best_score = score
            best_row   = row

    return best_row if best_score >= 2 else None


def _find_column_indices(header_row):
    """
    Build a field → column-index mapping from a header row.
    Uses keyword matching; falls back to default positional indices.
    """
    defaults = {
        'row_number':     0,
        'name':           1,
        'group':          2,
        'date':           3,
        'shift':          None,
        'engine_time':    4,
        'engine_no_move': 6,
        'engine_idle':    8,
        'fuel_norm':      10,
        'fuel_actual':    11,
        'downtime':       None,
        'mileage':        None,
        'refueling':      None,
        'dumptruck_norm': None,
    }

    if not header_row or all(c is None for c in header_row):
        return defaults

    header_lower = [str(c).lower().strip() if c else '' for c in header_row]

    # Keywords listed from most specific to least specific
    keywords = {
        'name':           ['транспортное средство', 'транспортное'],
        'group':          ['группа тс', 'группа'],
        'date':           ['дата'],
        'shift':          ['№ смены', 'номер смены', 'смена', 'shift'],
        'engine_time':    ['время работы двигателя', 'работы двигателя', 'работа двигателя'],
        'engine_no_move': ['без движения'],
        'engine_idle':    ['холостом ходу', 'на холостом', 'холостой ход', 'холостого хода'],
        'fuel_norm':      ['норма расхода', 'норм. расход'],
        'fuel_actual':    ['фактический расход', 'расход факт', 'факт. расход', 'факт расход'],
        'downtime':       ['простой стрелы', 'время простоя стрелы', 'время простоя'],
        'mileage':        ['пробег'],
        'refueling':      ['заправк', 'объём заправ', 'объем заправ'],
        'row_number':     ['№ п/п', '№ пп', '№'],
        'dumptruck_norm': ['норма самосвала', 'норма сам.', 'норма сам'],
    }

    # Fields that must NOT match headers containing these substrings
    exclusions = {
        'engine_time': ['холостом', 'без движения'],
    }

    for field, kws in keywords.items():
        exclude = exclusions.get(field, [])
        for kw in kws:
            for i, h in enumerate(header_lower):
                if kw in h and not any(ex in h for ex in exclude):
                    defaults[field] = i
                    break
            else:
                continue
            break   # found for this field, move on

    return defaults


# ─── Main parser ──────────────────────────────────────────────────────────────

def parse_excel_file(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    metadata = {'period': '', 'vehicles_list': '', 'report_name': '', 'year': None}
    records  = []

    # Read Omnicomm-style metadata from top rows
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        if not row or row[0] is None:
            continue
        first = str(row[0]).strip()
        second = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        if first == 'Отчет:':
            metadata['report_name'] = second
        elif first == 'Период:':
            metadata['period'] = second
        elif first == 'Транспортные средства:':
            metadata['vehicles_list'] = second

    # Detect header row (the row containing column names)
    header_row = _detect_header_row(ws)
    cols = _find_column_indices(header_row)

    # Determine start row (one after the header, or default 10)
    data_start = 10
    if header_row is not None:
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
            if row == header_row:
                data_start = i + 1
                break

    detected_year = None

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or row[0] is None:
            continue

        def get(idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        # ── Date ──
        date_str, record_date, row_year = parse_date_value(get(cols['date']))
        if row_year and detected_year is None:
            detected_year = row_year

        # ── Times ──
        engine_time_sec    = parse_timedelta_to_seconds(get(cols['engine_time']))    or 0
        engine_no_move_sec = parse_timedelta_to_seconds(get(cols['engine_no_move'])) or 0
        engine_idle_sec    = parse_timedelta_to_seconds(get(cols['engine_idle']))    or 0
        downtime_sec       = parse_timedelta_to_seconds(get(cols['downtime']))

        # ── Fuel ──
        fuel_norm   = parse_float(get(cols['fuel_norm']))   or 0
        fuel_actual = parse_float(get(cols['fuel_actual']))

        # ── Optional ──
        mileage   = parse_float(get(cols['mileage']))
        refueling = parse_float(get(cols['refueling']))

        # ── Dump-truck per-shift norm (HH:MM:SS timedelta or float hours) ──
        dumptruck_norm_sec = None
        raw_dt_norm = get(cols.get('dumptruck_norm'))
        if raw_dt_norm is not None and raw_dt_norm != '-' and raw_dt_norm != '':
            dumptruck_norm_sec = parse_timedelta_to_seconds(raw_dt_norm)
            # If it's a plain number, treat as hours
            if dumptruck_norm_sec is None:
                hours = parse_float(raw_dt_norm)
                if hours is not None and hours > 0:
                    dumptruck_norm_sec = hours * 3600

        # ── Row number ──
        try:
            row_num = int(float(str(get(cols['row_number'])).strip()))
        except (ValueError, TypeError):
            row_num = 0

        # ── Shift ──
        shift_val = 0
        if cols.get('shift') is not None:
            raw = get(cols['shift'])
            if raw is not None:
                try:
                    shift_val = int(float(str(raw).strip()))
                except (ValueError, TypeError):
                    shift_val = 0

        # ── Name / Group ──
        raw_name  = get(cols['name'])
        raw_group = get(cols['group'])
        name  = str(raw_name).strip()  if raw_name  else ''
        # Excel exports are inconsistent about casing ('самосвалы' vs 'Самосвалы'),
        # which broke type_efficiency matching downstream — normalize to the
        # capitalized form used everywhere else in the app.
        group = str(raw_group).strip().capitalize() if raw_group else ''

        # Skip rows that look like headers or totals (no meaningful name)
        if not name or name.lower() in ('транспортное средство', 'итого', 'всего', ''):
            continue

        records.append({
            'row_number':        row_num,
            'name':              name,
            'group':             group,
            'date':              date_str,
            'record_date':       record_date,
            'shift':             shift_val,
            'engine_time_sec':   engine_time_sec,
            'engine_no_move_sec': engine_no_move_sec,
            'engine_idle_sec':   engine_idle_sec,
            'fuel_norm':         fuel_norm,
            'fuel_actual':       fuel_actual,
            'downtime_sec':      downtime_sec,
            'mileage':           mileage,
            'refueling':         refueling,
            'dumptruck_norm_sec': dumptruck_norm_sec,
        })

    if detected_year:
        metadata['year'] = detected_year

    return metadata, records


# ─── Summary ──────────────────────────────────────────────────────────────────

def build_summary(vehicle_records, report):
    from collections import defaultdict

    groups = defaultdict(lambda: {
        'count': 0, 'anomaly_count': 0,
        'total_engine_hours': 0, 'total_fuel_actual': 0,
        'fuel_eff_sum': 0, 'fuel_eff_count': 0,
        'output_sum': 0, 'output_count': 0,
        'type_eff_sum': 0, 'type_eff_count': 0,
    })

    for rec in vehicle_records:
        g = groups[rec.group]
        g['count'] += 1
        if rec.has_anomaly:
            g['anomaly_count'] += 1
        g['total_engine_hours'] += rec.engine_time_sec / 3600
        if rec.fuel_actual is not None and rec.fuel_actual > 0:
            g['total_fuel_actual'] += rec.fuel_actual
        if rec.fuel_efficiency is not None:
            g['fuel_eff_sum']   += rec.fuel_efficiency
            g['fuel_eff_count'] += 1
        if rec.equipment_output is not None:
            g['output_sum']   += rec.equipment_output
            g['output_count'] += 1
        if rec.type_efficiency is not None:
            g['type_eff_sum']   += rec.type_efficiency
            g['type_eff_count'] += 1

    summary = []
    for group_name, data in sorted(groups.items()):
        avg_fuel_eff = (data['fuel_eff_sum'] / data['fuel_eff_count'] * 100) if data['fuel_eff_count'] > 0 else None
        avg_output   = (data['output_sum']   / data['output_count']   * 100) if data['output_count']   > 0 else None
        avg_type_eff = (data['type_eff_sum'] / data['type_eff_count'] * 100) if data['type_eff_count'] > 0 else None

        type_eff_label = {
            'Бульдозеры': f'Холостой ход к норме ({report.bulldozer_norm_str()})',
            'Погрузчики': f'Холостой ход к норме ({report.bulldozer_norm_str()})',
            'Экскаваторы': f'Простой стрелы к норме ({report.excavator_norm_str()})',
            'Самосвалы': f'Без движения к норме ({report.dumptruck_norm_str()})',
        }.get(group_name, 'Эффективность')

        summary.append({
            'group':            group_name,
            'count':            data['count'],
            'anomaly_count':    data['anomaly_count'],
            'total_engine_hours': round(data['total_engine_hours'], 1),
            'total_fuel_actual': round(data['total_fuel_actual'], 1),
            'avg_fuel_eff':     round(avg_fuel_eff, 1) if avg_fuel_eff is not None else None,
            'avg_output':       round(avg_output,   1) if avg_output   is not None else None,
            'avg_type_eff':     round(avg_type_eff, 1) if avg_type_eff is not None else None,
            'type_eff_label':   type_eff_label,
        })

    return summary
