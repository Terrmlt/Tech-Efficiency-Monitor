import io
import re
import json
import datetime
from functools import wraps
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db.models import Q, Avg, Count, Sum

from .forms import ReportUploadForm, SectionForm, UserCreateForm, UserEditForm
from .models import Report, VehicleRecord, Section, VehicleNorm, secs_to_hhmmss, UserProfile
from .utils import parse_excel_file, detect_anomalies, calculate_metrics, build_summary


# ─── Role helpers & decorators ────────────────────────────────────────────────

GROUP_ANALYST = 'Аналитика'
GROUP_MONITOR = 'Мониторинг'


def is_monitor(user):
    return user.groups.filter(name=GROUP_MONITOR).exists()


def is_analyst(user):
    return user.groups.filter(name=GROUP_ANALYST).exists()


def _get_user_section(user):
    try:
        return user.profile.section
    except Exception:
        return None


def role_required(*allowed_groups):
    """Allow access if user.is_staff OR user is in any of allowed_groups."""
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect(f'/login/?next={request.path}')
            if request.user.is_staff:
                return view_func(request, *args, **kwargs)
            if any(request.user.groups.filter(name=g).exists() for g in allowed_groups):
                return view_func(request, *args, **kwargs)
            if is_monitor(request.user):
                return redirect('monitoring_index')
            if is_analyst(request.user):
                return redirect('analytics')
            return redirect('login')
        return _wrapped
    return decorator


def staff_required(view_func):
    """Allow access only to staff/admin users."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect(f'/login/?next={request.path}')
        if request.user.is_staff:
            return view_func(request, *args, **kwargs)
        if is_monitor(request.user):
            return redirect('monitoring_index')
        if is_analyst(request.user):
            return redirect('analytics')
        return redirect('login')
    return _wrapped


# ─── Index ────────────────────────────────────────────────────────────────────

@login_required
def index(request):
    user = request.user
    if not user.is_staff:
        if is_monitor(user):
            return redirect('monitoring_index')
        if is_analyst(user):
            return redirect('analytics')
    reports = Report.objects.select_related('section').all()
    return render(request, 'analysis/index.html', {'reports': reports})


# ─── Upload ───────────────────────────────────────────────────────────────────

@staff_required
def upload(request):
    if request.method == 'POST':
        form = ReportUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    report = form.save()
                    metadata, records_data = parse_excel_file(report.file.path)

                    if metadata.get('period'):
                        report.period = metadata['period']
                    if metadata.get('vehicles_list'):
                        report.vehicles_list = metadata['vehicles_list']
                    if metadata.get('report_name') and not report.name:
                        report.name = metadata['report_name']
                    if metadata.get('year'):
                        report.year = metadata['year']
                    report.save()

                    # Collect dump-truck norms from Excel for auto-fill
                    excel_norms = {}  # (vehicle_name, shift, date) -> norm_sec

                    for rec_data in records_data:
                        has_anomaly, anomaly_details = detect_anomalies(rec_data)
                        metrics = calculate_metrics(rec_data, report)

                        VehicleRecord.objects.create(
                            report=report,
                            row_number=rec_data['row_number'],
                            name=rec_data['name'],
                            group=rec_data['group'],
                            date=rec_data['date'],
                            record_date=rec_data.get('record_date'),
                            shift=rec_data.get('shift', 0),
                            engine_time_sec=rec_data['engine_time_sec'],
                            engine_no_move_sec=rec_data['engine_no_move_sec'],
                            engine_idle_sec=rec_data['engine_idle_sec'],
                            fuel_norm=rec_data['fuel_norm'],
                            fuel_actual=rec_data['fuel_actual'],
                            downtime_sec=rec_data['downtime_sec'],
                            mileage=rec_data.get('mileage'),
                            refueling=rec_data.get('refueling'),
                            has_anomaly=has_anomaly,
                            anomaly_details=anomaly_details,
                            fuel_efficiency=metrics['fuel_efficiency'],
                            equipment_output=metrics['equipment_output'],
                            type_efficiency=metrics['type_efficiency'],
                        )

                        # Collect dump-truck per-shift norms from Excel column
                        dt_norm = rec_data.get('dumptruck_norm_sec')
                        if dt_norm and rec_data.get('group') == 'Самосвалы':
                            key = (rec_data['name'], rec_data.get('shift') or None, rec_data['date'])
                            excel_norms[key] = int(dt_norm)

                    # Auto-create VehicleNorm entries from Excel norms
                    for (vehicle_name, shift, date), norm_sec in excel_norms.items():
                        VehicleNorm.objects.update_or_create(
                            report=report,
                            vehicle_name=vehicle_name,
                            shift=shift,
                            date=date,
                            defaults={'dumptruck_norm_sec': norm_sec},
                        )

                    # Recalculate type_efficiency for dump trucks using the saved norms
                    if excel_norms:
                        norm_map = {
                            (vn.vehicle_name, vn.shift, vn.date): vn.dumptruck_norm_sec
                            for vn in VehicleNorm.objects.filter(report=report)
                        }
                        for rec in report.vehiclerecord_set.filter(group='Самосвалы'):
                            norm_sec = norm_map.get((rec.name, rec.shift if rec.shift else None, rec.date))
                            if norm_sec and norm_sec > 0:
                                rec.type_efficiency = rec.engine_no_move_sec / norm_sec
                                rec.save(update_fields=['type_efficiency'])

                messages.success(request, f'Отчёт успешно загружен: {report.vehiclerecord_set.count()} записей.')
                return redirect('report_detail', pk=report.pk)

            except Exception as e:
                messages.error(request, f'Ошибка при обработке файла: {str(e)}')
                if 'report' in locals() and report.pk:
                    report.delete()
    else:
        form = ReportUploadForm()

    return render(request, 'analysis/upload.html', {'form': form})


# ─── Daily-view helper ────────────────────────────────────────────────────────

def _build_daily_view(records, report):
    """
    Groups records by (name, date).  Returns a flat list of row-dicts:
      {'type': 'record',      'obj': VehicleRecord}
      {'type': 'daily_total', ...computed display fields...}
    Total rows are inserted after each group that has >1 shift entry.
    """
    from .models import secs_to_hhmmss

    order = []
    groups = {}
    for rec in records:
        key = (rec.name, rec.date)
        if key not in groups:
            order.append(key)
            groups[key] = []
        groups[key].append(rec)

    rows = []
    for key in order:
        name, date = key
        recs = sorted(groups[key], key=lambda r: r.shift if r.shift else 99)

        for rec in recs:
            ov = 0
            if rec.group in ('Бульдозеры', 'Погрузчики') and report.bulldozer_norm_sec > 0:
                ov = (rec.engine_idle_sec or 0) - report.bulldozer_norm_sec
            elif rec.group == 'Экскаваторы' and report.excavator_norm_sec > 0 and rec.downtime_sec is not None:
                ov = rec.downtime_sec - report.excavator_norm_sec
            # Dump trucks: over_str is set later by report_detail enrichment loop (uses VehicleNorm)
            rows.append({'type': 'record', 'obj': rec, 'over_str': secs_to_hhmmss(ov) if ov > 0 else ''})

        if len(recs) > 1:
            n = len(recs)
            total_engine   = sum(r.engine_time_sec for r in recs)
            total_no_move  = sum(r.engine_no_move_sec for r in recs)
            total_idle     = sum(r.engine_idle_sec for r in recs)
            has_fuel       = any(r.fuel_actual is not None for r in recs)
            total_fuel     = sum(r.fuel_actual for r in recs if r.fuel_actual is not None) if has_fuel else None
            has_mileage    = any(r.mileage is not None for r in recs)
            total_mileage  = sum(r.mileage for r in recs if r.mileage is not None) if has_mileage else None
            has_ref        = any(r.refueling is not None for r in recs)
            total_ref      = sum(r.refueling for r in recs if r.refueling is not None) if has_ref else None
            has_dt         = any(r.downtime_sec is not None for r in recs)
            total_downtime = sum(r.downtime_sec for r in recs if r.downtime_sec is not None) if has_dt else None

            fuel_norm = recs[0].fuel_norm
            group     = recs[0].group

            daily_norm_n  = report.daily_norm_sec * n
            daily_hours_n = daily_norm_n / 3600 if daily_norm_n > 0 else 0

            fuel_eff = None
            if fuel_norm > 0 and daily_hours_n > 0 and total_fuel is not None and total_fuel >= 0:
                fuel_eff = round(total_fuel / (fuel_norm * daily_hours_n) * 100, 1)

            output = round(total_engine / daily_norm_n * 100, 1) if daily_norm_n > 0 else None

            type_eff = None
            if group in ('Бульдозеры', 'Погрузчики') and report.bulldozer_norm_sec > 0:
                type_eff = round(total_idle / (report.bulldozer_norm_sec * n) * 100, 1)
            elif group == 'Экскаваторы' and total_downtime is not None and report.excavator_norm_sec > 0:
                type_eff = round(total_downtime / (report.excavator_norm_sec * n) * 100, 1)
            # Dump truck daily_total efficiency is computed later in report_detail
            # after per-shift VehicleNorm data is loaded.

            rows.append({
                'type':                 'daily_total',
                'name':                 name,
                'date':                 date,
                'group':                group,
                'n':                    n,
                'has_anomaly':          any(r.has_anomaly for r in recs),
                'engine_time_str':      secs_to_hhmmss(total_engine),
                'engine_idle_str':      secs_to_hhmmss(total_idle),
                'engine_idle_sec':      total_idle,
                'engine_no_move_str':   secs_to_hhmmss(total_no_move),
                'engine_no_move_sec':   total_no_move,
                'downtime_str':         secs_to_hhmmss(total_downtime) if total_downtime is not None else '—',
                'downtime_sec':         total_downtime,
                'fuel_actual':          total_fuel,
                'fuel_norm':            fuel_norm,
                'mileage':              total_mileage,
                'refueling':            total_ref,
                'fuel_efficiency_pct':  fuel_eff,
                'equipment_output_pct': output,
                'type_efficiency_pct':  type_eff,
                'is_bulldozer_or_loader': group in ('Бульдозеры', 'Погрузчики'),
                'is_excavator':         group == 'Экскаваторы',
                'is_dumptruck':         group == 'Самосвалы',
            })

    return rows


# ─── Report detail ────────────────────────────────────────────────────────────

@staff_required
def report_detail(request, pk):
    report = get_object_or_404(Report, pk=pk)
    all_records = report.vehiclerecord_set.all()

    group_filter  = request.GET.get('group', '')
    anomaly_filter = request.GET.get('anomaly', '')
    shift_filter  = request.GET.get('shift', '')

    filtered = all_records
    if group_filter:
        filtered = filtered.filter(group=group_filter)
    if anomaly_filter == 'yes':
        filtered = filtered.filter(has_anomaly=True)
    elif anomaly_filter == 'no':
        filtered = filtered.filter(has_anomaly=False)
    if shift_filter:
        try:
            filtered = filtered.filter(shift=int(shift_filter))
        except ValueError:
            pass

    summary    = build_summary(all_records, report)
    groups     = all_records.values_list('group', flat=True).distinct().order_by('group')
    daily_view = _build_daily_view(filtered.order_by('row_number', 'shift'), report)

    # Per-vehicle-per-shift-per-date norms map for dump trucks
    # Key: (vehicle_name, shift, date)
    existing_norms = {
        (vn.vehicle_name, vn.shift, vn.date): vn
        for vn in VehicleNorm.objects.filter(report=report)
    }
    # Enrich rows with norm string, efficiency %, and overage time for all groups
    for row in daily_view:
        if row['type'] == 'record':
            rec = row['obj']
            if rec.group == 'Самосвалы':
                vn = existing_norms.get((rec.name, rec.shift, rec.date))
                if vn and vn.dumptruck_norm_sec:
                    row['dt_norm_str'] = vn.norm_str()
                    overage_sec = (rec.engine_no_move_sec or 0) - vn.dumptruck_norm_sec
                    row['over_str'] = secs_to_hhmmss(overage_sec) if overage_sec > 0 else ''
                else:
                    row['dt_norm_str'] = ''
                    row['over_str'] = ''
            elif rec.group in ('Бульдозеры', 'Погрузчики'):
                if report.bulldozer_norm_sec > 0:
                    overage_sec = (rec.engine_idle_sec or 0) - report.bulldozer_norm_sec
                    row['over_str'] = secs_to_hhmmss(overage_sec) if overage_sec > 0 else ''
                else:
                    row['over_str'] = ''
            elif rec.group == 'Экскаваторы':
                if report.excavator_norm_sec > 0 and rec.downtime_sec is not None:
                    overage_sec = rec.downtime_sec - report.excavator_norm_sec
                    row['over_str'] = secs_to_hhmmss(overage_sec) if overage_sec > 0 else ''
                else:
                    row['over_str'] = ''
        elif row['type'] == 'daily_total' and row.get('is_dumptruck'):
            row_date = row.get('date', '')
            total_norm_sec = 0
            for shift_key in [1, 2]:
                vn = existing_norms.get((row['name'], shift_key, row_date))
                if vn and vn.dumptruck_norm_sec:
                    total_norm_sec += vn.dumptruck_norm_sec
            if total_norm_sec:
                row['dt_norm_str'] = secs_to_hhmmss(total_norm_sec)
                no_move_sec = row.get('engine_no_move_sec') or 0
                row['type_efficiency_pct'] = round(no_move_sec / total_norm_sec * 100, 1)
                overage_sec = no_move_sec - total_norm_sec
                row['over_str'] = secs_to_hhmmss(overage_sec) if overage_sec > 0 else ''
            else:
                row['dt_norm_str'] = ''
                row['over_str'] = ''
        elif row['type'] == 'daily_total' and row.get('is_bulldozer_or_loader'):
            if report.bulldozer_norm_sec > 0:
                idle_sec = row.get('engine_idle_sec') or 0
                n_shifts = row.get('n', 1)
                overage_sec = idle_sec - report.bulldozer_norm_sec * n_shifts
                row['over_str'] = secs_to_hhmmss(overage_sec) if overage_sec > 0 else ''
            else:
                row['over_str'] = ''
        elif row['type'] == 'daily_total' and row.get('is_excavator'):
            if report.excavator_norm_sec > 0:
                dt_sec = row.get('downtime_sec') or 0
                n_shifts = row.get('n', 1)
                overage_sec = dt_sec - report.excavator_norm_sec * n_shifts
                row['over_str'] = secs_to_hhmmss(overage_sec) if overage_sec > 0 else ''
            else:
                row['over_str'] = ''

    context = {
        'report':         report,
        'daily_view':     daily_view,
        'summary':        summary,
        'groups':         groups,
        'group_filter':   group_filter,
        'anomaly_filter': anomaly_filter,
        'shift_filter':   shift_filter,
        'total_count':    all_records.count(),
        'anomaly_count':  all_records.filter(has_anomaly=True).count(),
        'shown_count':    filtered.count(),
    }
    return render(request, 'analysis/report_detail.html', context)


# ─── Vehicle norms (dump trucks) ──────────────────────────────────────────────

def _parse_hhmmss_to_sec(value):
    """Parse 'HH:MM:SS' or 'HH:MM' string to total seconds. Returns None on error."""
    if not value:
        return None
    parts = value.strip().split(':')
    try:
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(float(parts[2]))
        if len(parts) == 2:
            return int(parts[0]) * 3600 + int(parts[1]) * 60
    except (ValueError, IndexError):
        return None
    return None


@staff_required
@require_POST
def set_vehicle_norms(request, pk):
    """Save per-day per-shift dump truck norms and recalculate efficiency.

    Expects JSON body:
      {"norms": [{"vehicle_name": "...", "shift": 1, "date": "19.06", "norm_str": "06:10:00"}, ...]}
    Each (vehicle_name, shift, date) tuple gets its own norm.
    """
    report = get_object_or_404(Report, pk=pk)
    try:
        data = json.loads(request.body)
        norms_list = data.get('norms', [])
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Неверный формат данных'}, status=400)

    updated_records = 0
    errors = []

    with transaction.atomic():
        for item in norms_list:
            vehicle_name = (item.get('vehicle_name') or '').strip()
            shift = item.get('shift')  # int or None
            date_str = (item.get('date') or '').strip()
            norm_str = (item.get('norm_str') or '').strip()

            if not vehicle_name or not norm_str:
                continue

            norm_sec = _parse_hhmmss_to_sec(norm_str)
            if norm_sec is None or norm_sec <= 0:
                errors.append(f'{vehicle_name} С{shift} {date_str}: неверный формат нормы «{norm_str}»')
                continue

            # Save norm for this specific (vehicle, shift, date) combination.
            # shift=None safe path: use filter+update/create to avoid SQLite NULL uniqueness issue.
            if shift is not None:
                VehicleNorm.objects.update_or_create(
                    report=report,
                    vehicle_name=vehicle_name,
                    shift=shift,
                    date=date_str,
                    defaults={'dumptruck_norm_sec': norm_sec},
                )
            else:
                updated_count = VehicleNorm.objects.filter(
                    report=report, vehicle_name=vehicle_name,
                    shift__isnull=True, date=date_str,
                ).update(dumptruck_norm_sec=norm_sec)
                if updated_count == 0:
                    VehicleNorm.objects.create(
                        report=report, vehicle_name=vehicle_name,
                        shift=None, date=date_str, dumptruck_norm_sec=norm_sec,
                    )

            # Recalculate type_efficiency only for the matching record(s)
            rec_qs = VehicleRecord.objects.filter(
                report=report, name=vehicle_name, group='Самосвалы',
            )
            if shift is not None:
                rec_qs = rec_qs.filter(shift=shift)
            if date_str:
                rec_qs = rec_qs.filter(date=date_str)

            for rec in rec_qs:
                rec.type_efficiency = rec.engine_no_move_sec / norm_sec
                rec.save(update_fields=['type_efficiency'])
                updated_records += 1

    return JsonResponse({
        'ok': True,
        'updated': updated_records,
        'errors': errors,
    })


# ─── Delete report ────────────────────────────────────────────────────────────

@staff_required
def delete_report(request, pk):
    report = get_object_or_404(Report, pk=pk)
    if request.method == 'POST':
        name = report.name
        report.delete()
        messages.success(request, f'Отчёт «{name}» удалён.')
    return redirect('index')


# ─── Save comment (AJAX) ──────────────────────────────────────────────────────

@staff_required
@require_POST
def save_comment(request, pk):
    rec = get_object_or_404(VehicleRecord, pk=pk)
    try:
        data = json.loads(request.body)
        rec.comment = data.get('comment', '').strip()
        rec.save(update_fields=['comment'])
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


# ─── Unified records view ─────────────────────────────────────────────────────

def _build_daily_view_records(records):
    """
    Like _build_daily_view but works across multiple reports.
    Groups by (name, date, report_id) so records from different reports
    for the same vehicle are kept separate.
    Each group's norms come from its own report object.
    Per-day per-shift dump truck norms are loaded from VehicleNorm.
    """
    from .models import secs_to_hhmmss

    # Collect all report IDs present in this queryset
    report_ids = set()
    rec_list = list(records)
    for rec in rec_list:
        report_ids.add(rec.report_id)

    # Load per-day per-shift norms for dump trucks: (report_id, vehicle_name, shift, date)
    dt_norms = {}
    for vn in VehicleNorm.objects.filter(report_id__in=report_ids):
        if vn.dumptruck_norm_sec:
            dt_norms[(vn.report_id, vn.vehicle_name, vn.shift, vn.date)] = vn.dumptruck_norm_sec

    order = []
    groups = {}
    for rec in rec_list:
        key = (rec.name, rec.date, rec.report_id)
        if key not in groups:
            order.append(key)
            groups[key] = []
        groups[key].append(rec)

    rows = []
    for key in order:
        name, date, report_id = key
        recs = sorted(groups[key], key=lambda r: r.shift if r.shift else 99)
        report = recs[0].report

        for rec in recs:
            ov = 0
            if rec.group in ('Бульдозеры', 'Погрузчики') and report.bulldozer_norm_sec > 0:
                ov = (rec.engine_idle_sec or 0) - report.bulldozer_norm_sec
            elif rec.group == 'Экскаваторы' and report.excavator_norm_sec > 0 and rec.downtime_sec is not None:
                ov = rec.downtime_sec - report.excavator_norm_sec
            elif rec.group == 'Самосвалы':
                # Use per-day per-shift VehicleNorm only — no global fallback
                vn_sec = dt_norms.get((report_id, rec.name, rec.shift, rec.date))
                if vn_sec:
                    ov = (rec.engine_no_move_sec or 0) - vn_sec
            rows.append({'type': 'record', 'obj': rec, 'over_str': secs_to_hhmmss(ov) if ov > 0 else ''})

        if len(recs) > 1:
            n = len(recs)
            total_engine   = sum(r.engine_time_sec for r in recs)
            total_no_move  = sum(r.engine_no_move_sec for r in recs)
            total_idle     = sum(r.engine_idle_sec for r in recs)
            has_fuel       = any(r.fuel_actual is not None for r in recs)
            total_fuel     = sum(r.fuel_actual for r in recs if r.fuel_actual is not None) if has_fuel else None
            has_mileage    = any(r.mileage is not None for r in recs)
            total_mileage  = sum(r.mileage for r in recs if r.mileage is not None) if has_mileage else None
            has_ref        = any(r.refueling is not None for r in recs)
            total_ref      = sum(r.refueling for r in recs if r.refueling is not None) if has_ref else None
            has_dt         = any(r.downtime_sec is not None for r in recs)
            total_downtime = sum(r.downtime_sec for r in recs if r.downtime_sec is not None) if has_dt else None

            fuel_norm = recs[0].fuel_norm
            group     = recs[0].group

            daily_norm_n  = report.daily_norm_sec * n
            daily_hours_n = daily_norm_n / 3600 if daily_norm_n > 0 else 0

            fuel_eff = None
            if fuel_norm > 0 and daily_hours_n > 0 and total_fuel is not None and total_fuel >= 0:
                fuel_eff = round(total_fuel / (fuel_norm * daily_hours_n) * 100, 1)

            output = round(total_engine / daily_norm_n * 100, 1) if daily_norm_n > 0 else None

            # For dump trucks: use per-day norms from VehicleNorm (sum of both shifts)
            dt_total_norm = 0
            if group == 'Самосвалы':
                for shift_key in [1, 2]:
                    s = dt_norms.get((report_id, name, shift_key, date))
                    if s:
                        dt_total_norm += s

            type_eff = None
            if group in ('Бульдозеры', 'Погрузчики') and report.bulldozer_norm_sec > 0:
                type_eff = round(total_idle / (report.bulldozer_norm_sec * n) * 100, 1)
            elif group == 'Экскаваторы' and total_downtime is not None and report.excavator_norm_sec > 0:
                type_eff = round(total_downtime / (report.excavator_norm_sec * n) * 100, 1)
            elif group == 'Самосвалы' and dt_total_norm > 0:
                type_eff = round(total_no_move / dt_total_norm * 100, 1)

            ov_total = 0
            if group in ('Бульдозеры', 'Погрузчики') and report.bulldozer_norm_sec > 0:
                ov_total = total_idle - report.bulldozer_norm_sec * n
            elif group == 'Экскаваторы' and total_downtime is not None and report.excavator_norm_sec > 0:
                ov_total = total_downtime - report.excavator_norm_sec * n
            elif group == 'Самосвалы' and dt_total_norm > 0:
                ov_total = total_no_move - dt_total_norm

            rows.append({
                'type':                 'daily_total',
                'name':                 name,
                'date':                 date,
                'record_date':          recs[0].record_date,
                'group':                group,
                'report':               report,
                'has_anomaly':          any(r.has_anomaly for r in recs),
                'engine_time_str':      secs_to_hhmmss(total_engine),
                'engine_idle_str':      secs_to_hhmmss(total_idle),
                'engine_idle_sec':      total_idle,
                'engine_no_move_str':   secs_to_hhmmss(total_no_move),
                'engine_no_move_sec':   total_no_move,
                'downtime_str':         secs_to_hhmmss(total_downtime) if total_downtime is not None else '—',
                'downtime_sec':         total_downtime,
                'fuel_actual':          total_fuel,
                'fuel_norm':            fuel_norm,
                'mileage':              total_mileage,
                'refueling':            total_ref,
                'fuel_efficiency_pct':  fuel_eff,
                'equipment_output_pct': output,
                'type_efficiency_pct':  type_eff,
                'over_str':             secs_to_hhmmss(ov_total) if ov_total > 0 else '',
                'is_bulldozer_or_loader': group in ('Бульдозеры', 'Погрузчики'),
                'is_excavator':         group == 'Экскаваторы',
                'is_dumptruck':         group == 'Самосвалы',
            })

    return rows


@staff_required
def records(request):
    qs = VehicleRecord.objects.select_related('report', 'report__section').all()

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    section_id = request.GET.get('section', '')
    group_filter = request.GET.get('group', '')
    anomaly_filter = request.GET.get('anomaly', '')
    shift_filter = request.GET.get('shift', '')

    if date_from:
        try:
            qs = qs.filter(record_date__gte=datetime.date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(record_date__lte=datetime.date.fromisoformat(date_to))
        except ValueError:
            pass
    if section_id:
        qs = qs.filter(report__section_id=section_id)
    if group_filter:
        qs = qs.filter(group=group_filter)
    if anomaly_filter == 'yes':
        qs = qs.filter(has_anomaly=True)
    elif anomaly_filter == 'no':
        qs = qs.filter(has_anomaly=False)
    if shift_filter:
        try:
            qs = qs.filter(shift=int(shift_filter))
        except ValueError:
            pass

    total_count = qs.count()
    anomaly_count = qs.filter(has_anomaly=True).count()

    ordered = qs.order_by('record_date', 'name', 'report_id', 'shift', 'row_number')
    daily_view = _build_daily_view_records(ordered)

    sections = Section.objects.all()
    groups = VehicleRecord.objects.values_list('group', flat=True).distinct().order_by('group')

    context = {
        'daily_view': daily_view,
        'sections': sections,
        'groups': groups,
        'date_from': date_from,
        'date_to': date_to,
        'section_id': section_id,
        'group_filter': group_filter,
        'anomaly_filter': anomaly_filter,
        'shift_filter': shift_filter,
        'total_count': total_count,
        'anomaly_count': anomaly_count,
    }
    return render(request, 'analysis/records.html', context)


# ─── Analytics ────────────────────────────────────────────────────────────────

@role_required(GROUP_ANALYST)
def analytics(request):
    from collections import defaultdict as _dd

    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')
    section_id = request.GET.get('section', '')
    group_filters = request.GET.getlist('group')

    qs = VehicleRecord.objects.select_related('report', 'report__section').all()

    if date_from:
        try:
            qs = qs.filter(record_date__gte=datetime.date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(record_date__lte=datetime.date.fromisoformat(date_to))
        except ValueError:
            pass
    if section_id:
        qs = qs.filter(report__section_id=section_id)
    if group_filters:
        qs = qs.filter(group__in=group_filters)

    # All available groups for checkboxes
    all_groups = list(VehicleRecord.objects.values_list('group', flat=True).distinct().order_by('group'))

    # ── helper ──────────────────────────────────────────────────────────────────
    def _avg(recs):
        fe = [r.fuel_efficiency * 100 for r in recs if r.fuel_efficiency is not None]
        ou = [r.equipment_output * 100 for r in recs if r.equipment_output is not None]
        te = [r.type_efficiency  * 100 for r in recs if r.type_efficiency  is not None]
        return (
            round(sum(fe) / len(fe), 1) if fe else None,
            round(sum(ou) / len(ou), 1) if ou else None,
            round(sum(te) / len(te), 1) if te else None,
        )

    # ── build daily indexes ──────────────────────────────────────────────────────
    daily_recs    = _dd(list)   # date -> [rec]
    vehicle_daily = _dd(lambda: _dd(list))  # vehicle -> date -> [rec]
    group_daily   = _dd(lambda: _dd(list))  # group   -> date -> [rec]

    rec_list = list(qs)
    for rec in rec_list:
        if rec.record_date:
            d = rec.record_date.isoformat()
            daily_recs[d].append(rec)
            vehicle_daily[rec.name][d].append(rec)
            group_daily[rec.group][d].append(rec)

    all_dates = sorted(daily_recs.keys())

    # ── overall daily ────────────────────────────────────────────────────────────
    ov_fe, ov_ou, ov_te = [], [], []
    for d in all_dates:
        fe, ou, te = _avg(daily_recs[d])
        ov_fe.append(fe); ov_ou.append(ou); ov_te.append(te)

    # ── by group daily ───────────────────────────────────────────────────────────
    by_group = {}
    for grp in sorted(group_daily.keys()):
        gfe, gou, gte = [], [], []
        for d in all_dates:
            recs = group_daily[grp].get(d, [])
            fe, ou, te = _avg(recs) if recs else (None, None, None)
            gfe.append(fe); gou.append(ou); gte.append(te)
        by_group[grp] = {'fuel_eff': gfe, 'output': gou, 'type_eff': gte}

    # ── by vehicle daily ─────────────────────────────────────────────────────────
    by_vehicle = {}
    for vname in sorted(vehicle_daily.keys()):
        date_map = vehicle_daily[vname]
        vfe, vou, vte, vgroup = [], [], [], None
        for d in all_dates:
            recs = date_map.get(d, [])
            if recs:
                if vgroup is None:
                    vgroup = recs[0].group
                fe, ou, te = _avg(recs)
            else:
                fe, ou, te = None, None, None
            vfe.append(fe); vou.append(ou); vte.append(te)
        by_vehicle[vname] = {'group': vgroup, 'fuel_eff': vfe, 'output': vou, 'type_eff': vte}

    trend_json = json.dumps({
        'dates':      all_dates,
        'overall':    {'fuel_eff': ov_fe,  'output': ov_ou,  'type_eff': ov_te},
        'by_group':   by_group,
        'by_vehicle': by_vehicle,
    })

    # ── group summary cards ──────────────────────────────────────────────────────
    groups_in_qs = list(qs.values_list('group', flat=True).distinct().order_by('group'))
    group_stats = []
    for group in groups_in_qs:
        recs = [r for r in rec_list if r.group == group]
        if not recs:
            continue
        total_fuel  = sum(r.fuel_actual for r in recs if r.fuel_actual and r.fuel_actual > 0)
        total_hours = sum(r.engine_time_sec for r in recs) / 3600
        fe_vals = [r.fuel_efficiency * 100 for r in recs if r.fuel_efficiency is not None]
        ou_vals = [r.equipment_output * 100 for r in recs if r.equipment_output is not None]
        te_vals = [r.type_efficiency  * 100 for r in recs if r.type_efficiency  is not None]
        group_stats.append({
            'group':        group,
            'count':        len(recs),
            'total_fuel':   round(total_fuel, 1),
            'total_hours':  round(total_hours, 1),
            'avg_fuel_eff': round(sum(fe_vals) / len(fe_vals), 1) if fe_vals else None,
            'avg_output':   round(sum(ou_vals) / len(ou_vals), 1) if ou_vals else None,
            'avg_type_eff': round(sum(te_vals) / len(te_vals), 1) if te_vals else None,
        })

    # ── totals ───────────────────────────────────────────────────────────────────
    total_count = len(rec_list)
    total_fuel  = sum(r.fuel_actual for r in rec_list if r.fuel_actual and r.fuel_actual > 0)
    total_hours = sum(r.engine_time_sec for r in rec_list) / 3600

    sections = Section.objects.all()

    context = {
        'group_stats':    group_stats,
        'trend_json':     trend_json,
        'has_trend':      bool(all_dates),
        'total_count':    total_count,
        'total_fuel':     round(total_fuel, 1),
        'total_hours':    round(total_hours, 1),
        'sections':       sections,
        'all_groups':     all_groups,
        'date_from':      date_from,
        'date_to':        date_to,
        'section_id':     section_id,
        'group_filters':  group_filters,
    }
    return render(request, 'analysis/analytics.html', context)


# ─── Analytics Compare ────────────────────────────────────────────────────────

@role_required(GROUP_ANALYST)
def analytics_compare(request):
    from collections import defaultdict as _dd

    selected_ids  = request.GET.getlist('sections')
    date_from     = request.GET.get('date_from', '')
    date_to       = request.GET.get('date_to', '')
    group_filters = request.GET.getlist('group')

    all_sections = list(Section.objects.all())
    all_groups   = list(VehicleRecord.objects.values_list('group', flat=True).distinct().order_by('group'))
    comparison_data = []

    def _avg(recs):
        fe = [r.fuel_efficiency * 100 for r in recs if r.fuel_efficiency is not None]
        ou = [r.equipment_output * 100 for r in recs if r.equipment_output is not None]
        te = [r.type_efficiency  * 100 for r in recs if r.type_efficiency  is not None]
        return (
            round(sum(fe) / len(fe), 1) if fe else None,
            round(sum(ou) / len(ou), 1) if ou else None,
            round(sum(te) / len(te), 1) if te else None,
        )

    # Collect all dates across all selected sections for alignment
    all_dates_set = set()

    section_recs_map = {}  # section.pk -> list[rec]
    if selected_ids:
        for section in all_sections:
            if str(section.pk) not in selected_ids:
                continue
            qs = VehicleRecord.objects.select_related('report').filter(report__section=section)
            if date_from:
                try:
                    qs = qs.filter(record_date__gte=datetime.date.fromisoformat(date_from))
                except ValueError:
                    pass
            if date_to:
                try:
                    qs = qs.filter(record_date__lte=datetime.date.fromisoformat(date_to))
                except ValueError:
                    pass
            if group_filters:
                qs = qs.filter(group__in=group_filters)
            recs = list(qs)
            section_recs_map[section.pk] = recs
            for r in recs:
                if r.record_date:
                    all_dates_set.add(r.record_date.isoformat())

    all_dates = sorted(all_dates_set)

    if selected_ids:
        for section in all_sections:
            if str(section.pk) not in selected_ids:
                continue
            recs = section_recs_map.get(section.pk, [])
            count       = len(recs)
            total_fuel  = round(sum(r.fuel_actual for r in recs if r.fuel_actual and r.fuel_actual > 0), 1)
            total_hours = round(sum(r.engine_time_sec for r in recs) / 3600, 1)
            fe_vals = [r.fuel_efficiency * 100 for r in recs if r.fuel_efficiency is not None]
            ou_vals = [r.equipment_output * 100 for r in recs if r.equipment_output is not None]
            te_vals = [r.type_efficiency  * 100 for r in recs if r.type_efficiency  is not None]

            # Per-group breakdown
            groups_map = _dd(list)
            for rec in recs:
                groups_map[rec.group].append(rec)
            group_stats = []
            for grp in sorted(groups_map):
                g = groups_map[grp]
                gfe = [r.fuel_efficiency * 100 for r in g if r.fuel_efficiency is not None]
                go  = [r.equipment_output * 100 for r in g if r.equipment_output is not None]
                gte = [r.type_efficiency  * 100 for r in g if r.type_efficiency  is not None]
                group_stats.append({
                    'group':        grp,
                    'count':        len(g),
                    'avg_fuel_eff': round(sum(gfe)/len(gfe), 1) if gfe else None,
                    'avg_output':   round(sum(go)/len(go), 1)   if go  else None,
                    'avg_type_eff': round(sum(gte)/len(gte), 1) if gte else None,
                })

            # Daily trend for this section
            daily_map = _dd(list)
            vehicle_daily = _dd(lambda: _dd(list))
            for rec in recs:
                if rec.record_date:
                    d = rec.record_date.isoformat()
                    daily_map[d].append(rec)
                    vehicle_daily[rec.name][d].append(rec)

            sec_ov_fe, sec_ov_ou, sec_ov_te = [], [], []
            for d in all_dates:
                r_d = daily_map.get(d, [])
                fe, ou, te = _avg(r_d) if r_d else (None, None, None)
                sec_ov_fe.append(fe); sec_ov_ou.append(ou); sec_ov_te.append(te)

            # Per vehicle daily
            sec_vehicles = {}
            for vname in sorted(vehicle_daily.keys()):
                vfe, vou, vte, vgroup = [], [], [], None
                for d in all_dates:
                    r_d = vehicle_daily[vname].get(d, [])
                    if r_d:
                        if vgroup is None:
                            vgroup = r_d[0].group
                        fe, ou, te = _avg(r_d)
                    else:
                        fe, ou, te = None, None, None
                    vfe.append(fe); vou.append(ou); vte.append(te)
                sec_vehicles[vname] = {'group': vgroup, 'fuel_eff': vfe, 'output': vou, 'type_eff': vte}

            comparison_data.append({
                'section':      section,
                'count':        count,
                'total_fuel':   total_fuel,
                'total_hours':  total_hours,
                'avg_fuel_eff': round(sum(fe_vals)/len(fe_vals), 1) if fe_vals else None,
                'avg_output':   round(sum(ou_vals)/len(ou_vals), 1) if ou_vals else None,
                'avg_type_eff': round(sum(te_vals)/len(te_vals), 1) if te_vals else None,
                'group_stats':  group_stats,
                'daily_overall': {'fuel_eff': sec_ov_fe, 'output': sec_ov_ou, 'type_eff': sec_ov_te},
                'daily_vehicles': sec_vehicles,
            })

    PALETTE = ['#0d6efd','#198754','#fd7e14','#6f42c1','#dc3545','#0dcaf0','#ffc107','#20c997']
    for i, d in enumerate(comparison_data):
        d['color'] = PALETTE[i % len(PALETTE)]

    compare_json = json.dumps({
        'dates': all_dates,
        'sections': [
            {
                'name':         d['section'].name,
                'color':        d['color'],
                'total_hours':  d['total_hours'],
                'total_fuel':   d['total_fuel'],
                'avg_fuel_eff': d['avg_fuel_eff'],
                'avg_output':   d['avg_output'],
                'avg_type_eff': d['avg_type_eff'],
                'overall':      d['daily_overall'],
                'vehicles':     d['daily_vehicles'],
            }
            for d in comparison_data
        ],
    }) if comparison_data else json.dumps({'dates': [], 'sections': []})

    context = {
        'all_sections':    all_sections,
        'all_groups':      all_groups,
        'selected_ids':    selected_ids,
        'comparison_data': comparison_data,
        'date_from':       date_from,
        'date_to':         date_to,
        'group_filters':   group_filters,
        'compare_json':    compare_json,
        'has_trend':       bool(all_dates),
    }
    return render(request, 'analysis/compare.html', context)


# ─── Sections CRUD ────────────────────────────────────────────────────────────

@staff_required
def sections(request):
    all_sections = Section.objects.annotate(report_count=Count('report'))
    return render(request, 'analysis/sections.html', {'sections': all_sections})


@staff_required
def section_create(request):
    form = SectionForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Участок добавлен.')
        return redirect('sections')
    return render(request, 'analysis/section_form.html', {'form': form, 'title': 'Добавить участок'})


@staff_required
def section_edit(request, pk):
    section = get_object_or_404(Section, pk=pk)
    form = SectionForm(request.POST or None, instance=section)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Участок обновлён.')
        return redirect('sections')
    return render(request, 'analysis/section_form.html', {'form': form, 'title': 'Изменить участок'})


@staff_required
def section_delete(request, pk):
    section = get_object_or_404(Section, pk=pk)
    if request.method == 'POST':
        section.delete()
        messages.success(request, 'Участок удалён.')
        return redirect('sections')
    return render(request, 'analysis/section_confirm_delete.html', {'section': section})


# ─── User management ──────────────────────────────────────────────────────────

@staff_required
def users_list(request):
    from django.contrib.auth.models import User
    users = User.objects.select_related('profile', 'profile__section').prefetch_related('groups').order_by('username')
    all_sections = Section.objects.all()
    return render(request, 'analysis/users.html', {'users': users, 'all_sections': all_sections})


@staff_required
def user_create(request):
    form = UserCreateForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        messages.success(request, f'Пользователь «{user.username}» создан.')
        return redirect('users_list')
    return render(request, 'analysis/user_form.html', {'form': form, 'title': 'Создать пользователя'})


@staff_required
def user_edit(request, pk):
    from django.contrib.auth.models import User
    user_obj = get_object_or_404(User, pk=pk)
    profile, _ = UserProfile.objects.get_or_create(user=user_obj)
    current_group = user_obj.groups.first()

    initial = {
        'first_name': user_obj.first_name,
        'last_name': user_obj.last_name,
        'group': current_group,
        'section': profile.section,
    }
    form = UserEditForm(request.POST or None, initial=initial)
    if request.method == 'POST' and form.is_valid():
        form.save(user_obj)
        messages.success(request, f'Пользователь «{user_obj.username}» обновлён.')
        return redirect('users_list')
    return render(request, 'analysis/user_form.html', {
        'form': form,
        'title': f'Редактировать: {user_obj.username}',
        'edit_user': user_obj,
    })


@staff_required
def user_delete(request, pk):
    from django.contrib.auth.models import User
    user_obj = get_object_or_404(User, pk=pk)
    if request.user.pk == user_obj.pk:
        messages.error(request, 'Нельзя удалить собственный аккаунт.')
        return redirect('users_list')
    if request.method == 'POST':
        username = user_obj.username
        user_obj.delete()
        messages.success(request, f'Пользователь «{username}» удалён.')
        return redirect('users_list')
    return render(request, 'analysis/user_confirm_delete.html', {'edit_user': user_obj})


@staff_required
@require_POST
def user_set_section(request, pk):
    """Inline section update from the users list (AJAX)."""
    from django.contrib.auth.models import User
    user_obj = get_object_or_404(User, pk=pk)
    section_id = request.POST.get('section_id', '').strip()
    profile, _ = UserProfile.objects.get_or_create(user=user_obj)
    if section_id:
        section = get_object_or_404(Section, pk=section_id)
        profile.section = section
    else:
        profile.section = None
    profile.save()
    return JsonResponse({'ok': True})


# ─── Export Excel ─────────────────────────────────────────────────────────────

@staff_required
def export_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    report = get_object_or_404(Report, pk=pk)
    all_records = list(report.vehiclerecord_set.all().order_by('group', 'row_number'))
    summary = build_summary(all_records, report)

    wb = _build_excel_workbook(report, all_records, summary)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = re.sub(r'[^\w\-]', '_', report.name)[:40]
    filename = f'analysis_{safe_name}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@staff_required
def export_records_excel(request):
    """Export filtered records — grouped by equipment type, shifts hierarchically."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict, OrderedDict

    qs = VehicleRecord.objects.select_related('report', 'report__section').all()

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    section_id = request.GET.get('section', '')
    group_filter = request.GET.get('group', '')
    anomaly_filter = request.GET.get('anomaly', '')
    shift_filter = request.GET.get('shift', '')

    if date_from:
        try:
            qs = qs.filter(record_date__gte=datetime.date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(record_date__lte=datetime.date.fromisoformat(date_to))
        except ValueError:
            pass
    if section_id:
        qs = qs.filter(report__section_id=section_id)
    if group_filter:
        qs = qs.filter(group=group_filter)
    if anomaly_filter == 'yes':
        qs = qs.filter(has_anomaly=True)
    elif anomaly_filter == 'no':
        qs = qs.filter(has_anomaly=False)
    if shift_filter:
        try:
            qs = qs.filter(shift=int(shift_filter))
        except ValueError:
            pass

    all_records = list(qs.order_by('group', 'name', 'record_date', 'date', 'shift'))

    # Load per-day per-shift dump truck norms for all relevant reports
    report_ids = {rec.report_id for rec in all_records}
    export_dt_norms = {
        (vn.report_id, vn.vehicle_name, vn.shift, vn.date): vn.dumptruck_norm_sec
        for vn in VehicleNorm.objects.filter(report_id__in=report_ids)
        if vn.dumptruck_norm_sec
    }

    from .models import secs_to_hhmmss as _s2h

    # ── Style constants ────────────────────────────────────────────────
    DARK_BLUE = '1F3864'
    MID_BLUE  = '2E75B6'
    GROUP_BG  = 'D6E4F0'
    TOTAL_BG  = 'BDD7EE'
    YELLOW_BG = 'FFF2CC'
    GREEN_BG  = 'E2EFDA'
    RED_BG    = 'FFE0E0'
    ORANGE_BG = 'FCE4D6'

    NCOLS = 18

    def _side():
        return Side(style='thin', color='BFBFBF')

    def _border():
        s = _side()
        return Border(left=s, right=s, top=s, bottom=s)

    def _fill(color):
        return PatternFill('solid', fgColor=color)

    def _font(bold=False, color='000000', size=10):
        return Font(bold=bold, color=color, size=size, name='Calibri')

    def _align(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _pct_fill(val, good_max, warn_max):
        if val is None:
            return None
        if val <= good_max:
            return _fill(GREEN_BG)
        if val <= warn_max:
            return _fill(ORANGE_BG)
        return _fill(RED_BG)

    # ── Build workbook ─────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Записи'
    ws.sheet_properties.outlinePr.summaryBelow = True

    # ── Column headers (row 1) ─────────────────────────────────────────
    HEADERS = [
        'Дата', 'Смена', 'Участок', 'Отчёт', '№', 'Техника', 'Группа',
        'Время\nработы', 'Факт. расход\n(л)', 'Норма\n(л/ч)',
        'Расход к норме\n(%)', 'Выход техники\n(%)', 'Эффективность\n(%)',
        'Простой свыше\nнормы',
        'Пробег\n(км)', 'Заправка\n(л)', 'Аномалии', 'Комментарий',
    ]
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
        c.fill = _fill(MID_BLUE)
        c.alignment = _align(wrap=True)
        c.border = _border()
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = 'A2'

    # ── Helper: write one data row ─────────────────────────────────────
    def write_record_row(rec, outline_level=0):
        rpt = rec.report
        fuel_eff_pct  = round(rec.fuel_efficiency  * 100, 1) if rec.fuel_efficiency  is not None else None
        output_pct    = round(rec.equipment_output * 100, 1) if rec.equipment_output is not None else None
        type_eff_pct  = round(rec.type_efficiency  * 100, 1) if rec.type_efficiency  is not None else None

        ov = 0
        if rec.group in ('Бульдозеры', 'Погрузчики') and rpt.bulldozer_norm_sec > 0:
            ov = (rec.engine_idle_sec or 0) - rpt.bulldozer_norm_sec
        elif rec.group == 'Экскаваторы' and rpt.excavator_norm_sec > 0 and rec.downtime_sec is not None:
            ov = rec.downtime_sec - rpt.excavator_norm_sec
        elif rec.group == 'Самосвалы':
            vn_sec = export_dt_norms.get((rpt.pk, rec.name, rec.shift, rec.date))
            if vn_sec:
                ov = (rec.engine_no_move_sec or 0) - vn_sec
        over_str = _s2h(ov) if ov > 0 else ''

        section_name  = rpt.section.name if rpt.section else ''
        date_display  = rec.record_date.strftime('%d.%m.%Y') if rec.record_date else rec.date

        ws.append([
            date_display,
            f'С{rec.shift}' if rec.shift else '—',
            section_name, rpt.name, rec.row_number, rec.name, rec.group,
            rec.engine_time_str(), rec.fuel_actual, rec.fuel_norm,
            fuel_eff_pct, output_pct, type_eff_pct, over_str,
            rec.mileage, rec.refueling,
            '; '.join(rec.anomaly_details) if rec.has_anomaly else '',
            rec.comment,
        ])
        r = ws.max_row
        ws.row_dimensions[r].outline_level = outline_level
        base_fill = _fill(YELLOW_BG) if rec.has_anomaly else None
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=r, column=col)
            c.font = _font()
            c.alignment = _align()
            c.border = _border()
            if base_fill and col not in (11, 12, 13):
                c.fill = base_fill
        ws.cell(row=r, column=6).alignment  = _align('left')
        ws.cell(row=r, column=17).alignment = _align('left', wrap=True)
        ws.cell(row=r, column=18).alignment = _align('left', wrap=True)
        # Colour KPI cells
        f = _pct_fill(fuel_eff_pct, 100, 115)
        if f:
            ws.cell(row=r, column=11).fill = f
        if output_pct is not None:
            ws.cell(row=r, column=12).fill = _fill(GREEN_BG) if output_pct >= 80 else _fill(RED_BG)
        f = _pct_fill(type_eff_pct, 100, 130)
        if f:
            ws.cell(row=r, column=13).fill = f
        if over_str:
            ws.cell(row=r, column=14).fill = _fill('FFB3B3')
            ws.cell(row=r, column=14).font = Font(bold=True, color='C00000', name='Calibri', size=10)

    # ── Helper: write a daily-total row for multi-shift vehicle+date ───
    def write_total_row(recs, outline_level=0):
        rpt           = recs[0].report
        group         = recs[0].group
        date_display  = recs[0].record_date.strftime('%d.%m.%Y') if recs[0].record_date else recs[0].date
        section_name  = rpt.section.name if rpt.section else ''
        n             = len(recs)

        total_engine  = sum(r.engine_time_sec for r in recs)
        total_no_move = sum(r.engine_no_move_sec for r in recs)
        total_idle    = sum(r.engine_idle_sec for r in recs)
        total_fuel    = sum(r.fuel_actual for r in recs if r.fuel_actual is not None) if any(r.fuel_actual is not None for r in recs) else None
        total_mileage = sum(r.mileage for r in recs if r.mileage is not None) if any(r.mileage is not None for r in recs) else None
        total_ref     = sum(r.refueling for r in recs if r.refueling is not None) if any(r.refueling is not None for r in recs) else None
        total_dt      = sum(r.downtime_sec for r in recs if r.downtime_sec is not None) if any(r.downtime_sec is not None for r in recs) else None
        fuel_norm     = recs[0].fuel_norm

        daily_norm_n  = rpt.daily_norm_sec * n
        daily_h_n     = daily_norm_n / 3600 if daily_norm_n > 0 else 0

        fuel_eff_pct = None
        if fuel_norm > 0 and daily_h_n > 0 and total_fuel is not None and total_fuel >= 0:
            fuel_eff_pct = round(total_fuel / (fuel_norm * daily_h_n) * 100, 1)
        output_pct = round(total_engine / daily_norm_n * 100, 1) if daily_norm_n > 0 else None

        type_eff_pct = None
        if group in ('Бульдозеры', 'Погрузчики') and rpt.bulldozer_norm_sec > 0:
            type_eff_pct = round(total_idle / (rpt.bulldozer_norm_sec * n) * 100, 1)
        elif group == 'Экскаваторы' and total_dt is not None and rpt.excavator_norm_sec > 0:
            type_eff_pct = round(total_dt / (rpt.excavator_norm_sec * n) * 100, 1)
        elif group == 'Самосвалы':
            dt_total = sum(
                export_dt_norms.get((rpt.pk, recs[0].name, r.shift, r.date), 0)
                for r in recs
            )
            if dt_total > 0:
                type_eff_pct = round(total_no_move / dt_total * 100, 1)

        ov = 0
        if group in ('Бульдозеры', 'Погрузчики') and rpt.bulldozer_norm_sec > 0:
            ov = total_idle - rpt.bulldozer_norm_sec * n
        elif group == 'Экскаваторы' and total_dt is not None and rpt.excavator_norm_sec > 0:
            ov = total_dt - rpt.excavator_norm_sec * n
        over_str = _s2h(ov) if ov > 0 else ''

        anomaly_cnt = sum(1 for r in recs if r.has_anomaly)

        ws.append([
            date_display, 'Итого', section_name, rpt.name, '',
            recs[0].name, group,
            _s2h(total_engine), total_fuel, fuel_norm,
            fuel_eff_pct, output_pct, type_eff_pct, over_str,
            total_mileage, total_ref,
            f'Аномалий: {anomaly_cnt}' if anomaly_cnt else '', '',
        ])
        r = ws.max_row
        ws.row_dimensions[r].outline_level = outline_level
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=r, column=col)
            c.font = Font(bold=True, size=10, name='Calibri', color=DARK_BLUE)
            c.fill = _fill(TOTAL_BG)
            c.alignment = _align()
            c.border = _border()
        ws.cell(row=r, column=6).alignment  = _align('left')
        ws.cell(row=r, column=17).alignment = _align('left')
        f = _pct_fill(fuel_eff_pct, 100, 115)
        if f:
            ws.cell(row=r, column=11).fill = f
        if output_pct is not None:
            ws.cell(row=r, column=12).fill = _fill(GREEN_BG) if output_pct >= 80 else _fill(RED_BG)
        f = _pct_fill(type_eff_pct, 100, 130)
        if f:
            ws.cell(row=r, column=13).fill = f

    # ── Build hierarchy: group → vehicle → date → shifts ──────────────
    by_group = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for rec in all_records:
        by_group[rec.group][rec.name][rec.date].append(rec)

    for group_name in sorted(by_group.keys()):
        vehicles = by_group[group_name]
        total_recs = sum(
            len(shifts)
            for dates in vehicles.values()
            for shifts in dates.values()
        )

        # Group header
        ws.append([''] * NCOLS)
        gr = ws.max_row
        ws.merge_cells(f'A{gr}:{get_column_letter(NCOLS)}{gr}')
        gc = ws.cell(row=gr, column=1)
        gc.value = f'  {group_name.upper()}  ({total_recs} записей)'
        gc.font  = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
        gc.fill  = _fill(DARK_BLUE)
        gc.alignment = _align('left')
        ws.row_dimensions[gr].height = 20

        for vehicle_name in sorted(vehicles.keys()):
            dates = vehicles[vehicle_name]
            for date_key in sorted(dates.keys()):
                shifts = sorted(dates[date_key], key=lambda r: r.shift if r.shift else 99)
                multi  = len(shifts) > 1

                for rec in shifts:
                    write_record_row(rec, outline_level=1 if multi else 0)

                if multi:
                    write_total_row(shifts, outline_level=0)

    # ── Column widths ──────────────────────────────────────────────────
    col_widths = [12, 8, 16, 25, 5, 30, 14, 14, 14, 10, 14, 14, 14, 14, 12, 12, 45, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    period_str = f'{date_from}__{date_to}' if date_from or date_to else 'все'
    filename = f'records_{period_str}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _build_excel_workbook(report, all_records, summary):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    DARK_BLUE = '1F3864'
    MID_BLUE  = '2E75B6'
    YELLOW_BG = 'FFF2CC'
    ORANGE_BG = 'FCE4D6'
    GREEN_BG  = 'E2EFDA'
    RED_BG    = 'FFE0E0'
    GREY_BG   = 'F2F2F2'
    TOTAL_BG  = 'BDD7EE'

    def _fill(color):
        return PatternFill('solid', fgColor=color)

    def _border():
        s = Side(style='thin', color='BFBFBF')
        return Border(left=s, right=s, top=s, bottom=s)

    def _align(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _pct_fill(val, good_max, warn_max):
        if val is None:
            return None
        if val <= good_max:
            return _fill(GREEN_BG)
        if val <= warn_max:
            return _fill(YELLOW_BG)
        return _fill(ORANGE_BG)

    # Columns: 1=№ 2=Техника 3=Группа 4=Дата 5=Смена 6=Время работы
    # 7=Факт.расход 8=Норма(л/ч) 9=Расход% 10=Выход% 11=Эффект%
    # 12=Тип 13=Норма б/д 14=Превышение 15=Пробег 16=Заправка 17=Аномалии 18=Комментарий
    NCOLS = 18

    # Load dump-truck per-shift norms: (vehicle_name, shift, date) → seconds
    dt_norms = {
        (vn.vehicle_name, vn.shift, vn.date): vn.dumptruck_norm_sec
        for vn in VehicleNorm.objects.filter(report=report)
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Отчёт'
    ws.sheet_properties.outlinePr.summaryBelow = True

    # ── Title rows ─────────────────────────────────────────────────────
    def write_merged(row_num, text, bg, fg='FFFFFF', bold=True, size=11):
        ws.merge_cells(f'A{row_num}:{get_column_letter(NCOLS)}{row_num}')
        c = ws.cell(row=row_num, column=1)
        c.value = text
        c.font = Font(bold=bold, color=fg, size=size, name='Calibri')
        c.fill = _fill(bg)
        c.alignment = _align('center')
        ws.row_dimensions[row_num].height = 22

    write_merged(1, f'Анализ эффективности техники — {report.name}', DARK_BLUE, size=13)
    ws.row_dimensions[1].height = 28

    section_name = report.section.name if report.section else '—'
    write_merged(2, f'Период: {report.period} | Участок: {section_name}', 'E8F0FB', '595959', bold=False, size=10)

    norm_row = [
        'Норма/смену:', report.daily_norm_str(),
        'Хол.ход (бульдозеры):', report.bulldozer_norm_str(),
        'Простой стрелы (экскаваторы):', report.excavator_norm_str(),
    ] + [None] * (NCOLS - 6)
    ws.append(norm_row)
    for col in range(1, NCOLS + 1):
        c = ws.cell(row=3, column=col)
        c.font = Font(bold=(col % 2 == 1), size=9, name='Calibri', color='444444')
        c.fill = _fill(GREY_BG)
        c.alignment = _align('center')

    ws.append([])

    # ── Column headers (row 5) ─────────────────────────────────────────
    COL_HEADERS = [
        '№', 'Техника', 'Группа', 'Дата', 'Смена',
        'Время работы\n(чч:мм:сс)', 'Факт. расход\n(л)', 'Норма расхода\n(л/ч)',
        'Расход\nк норме (%)', 'Выход\nтехники (%)', 'Эффективность\n(%)',
        'Тип\nэффективности', 'Норма\nб/д (чч:мм:сс)', 'Превышение\nнормы б/д',
        'Пробег\n(км)', 'Заправка\n(л)', 'Аномалии', 'Комментарий',
    ]
    ws.append(COL_HEADERS)
    HDR_ROW = 5
    for col in range(1, NCOLS + 1):
        c = ws.cell(row=HDR_ROW, column=col)
        c.font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
        c.fill = _fill(MID_BLUE)
        c.alignment = _align(wrap=True)
        c.border = _border()
    ws.row_dimensions[HDR_ROW].height = 40
    ws.freeze_panes = f'A{HDR_ROW + 1}'

    TYPE_LABEL = {
        'Бульдозеры': 'Холостой ход',
        'Погрузчики': 'Холостой ход',
        'Экскаваторы': 'Простой стрелы',
        'Самосвалы': 'Без движения',
    }

    # ── Helper: style one data row ─────────────────────────────────────
    def _style_row(r, rec_has_anomaly, fuel_eff_pct, output_pct, type_eff_pct,
                   over_bd_str, fuel_actual, is_total=False):
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=r, column=col)
            c.font = Font(bold=is_total, size=10, name='Calibri',
                          color=DARK_BLUE if is_total else '000000')
            c.alignment = _align()
            c.border = _border()
            if is_total:
                c.fill = _fill(TOTAL_BG)
            elif rec_has_anomaly and col not in (9, 10, 11):
                c.fill = _fill(YELLOW_BG)
        ws.cell(row=r, column=2).alignment = _align('left')
        ws.cell(row=r, column=17).alignment = _align('left', wrap=True)
        ws.cell(row=r, column=18).alignment = _align('left', wrap=True)
        f = _pct_fill(fuel_eff_pct, 100, 115)
        if f:
            ws.cell(row=r, column=9).fill = f
        if output_pct is not None:
            ws.cell(row=r, column=10).fill = _fill(GREEN_BG) if output_pct >= 80 else _fill(RED_BG)
        f = _pct_fill(type_eff_pct, 100, 130)
        if f:
            ws.cell(row=r, column=11).fill = f
        if over_bd_str:
            ws.cell(row=r, column=14).fill = _fill('FFB3B3')
            ws.cell(row=r, column=14).font = Font(bold=True, color='C00000', name='Calibri', size=10)
        if fuel_actual is not None and fuel_actual < 0:
            ws.cell(row=r, column=7).fill = _fill('FFB3B3')
            ws.cell(row=r, column=7).font = Font(bold=True, color='C00000', name='Calibri', size=10)

    # ── Helper: write one shift row ────────────────────────────────────
    def write_rec_row(rec, outline_level=0):
        fuel_eff_pct = round(rec.fuel_efficiency  * 100, 1) if rec.fuel_efficiency  is not None else None
        output_pct   = round(rec.equipment_output * 100, 1) if rec.equipment_output is not None else None
        type_eff_pct = round(rec.type_efficiency  * 100, 1) if rec.type_efficiency  is not None else None

        norm_bd_str = over_bd_str = ''
        if rec.group == 'Самосвалы':
            norm_sec = dt_norms.get((rec.name, rec.shift, rec.date))
            if norm_sec:
                norm_bd_str = secs_to_hhmmss(norm_sec)
                ov = (rec.engine_no_move_sec or 0) - norm_sec
                if ov > 0:
                    over_bd_str = f'+{secs_to_hhmmss(ov)}'

        date_display  = rec.record_date.strftime('%d.%m.%Y') if rec.record_date else rec.date
        shift_display = f'С{rec.shift}' if rec.shift else '—'
        anomaly_text  = '; '.join(rec.anomaly_details) if rec.has_anomaly else ''

        ws.append([
            rec.row_number, rec.name, rec.group, date_display, shift_display,
            rec.engine_time_str(), rec.fuel_actual, rec.fuel_norm,
            fuel_eff_pct, output_pct, type_eff_pct,
            TYPE_LABEL.get(rec.group, ''),
            norm_bd_str, over_bd_str,
            rec.mileage, rec.refueling, anomaly_text, rec.comment,
        ])
        r = ws.max_row
        ws.row_dimensions[r].outline_level = outline_level
        _style_row(r, rec.has_anomaly, fuel_eff_pct, output_pct, type_eff_pct,
                   over_bd_str, rec.fuel_actual)
        return fuel_eff_pct, output_pct, type_eff_pct, rec

    # ── Helper: write a daily-total row for multi-shift vehicle+date ───
    def write_daily_total(recs, outline_level=0):
        n             = len(recs)
        group         = recs[0].group
        date_display  = recs[0].record_date.strftime('%d.%m.%Y') if recs[0].record_date else recs[0].date
        total_engine  = sum(r.engine_time_sec for r in recs)
        total_no_move = sum(r.engine_no_move_sec for r in recs)
        total_idle    = sum(r.engine_idle_sec for r in recs)
        total_fuel    = sum(r.fuel_actual for r in recs if r.fuel_actual is not None) if any(r.fuel_actual is not None for r in recs) else None
        total_mileage = sum(r.mileage for r in recs if r.mileage is not None) if any(r.mileage is not None for r in recs) else None
        total_ref     = sum(r.refueling for r in recs if r.refueling is not None) if any(r.refueling is not None for r in recs) else None
        total_dt      = sum(r.downtime_sec for r in recs if r.downtime_sec is not None) if any(r.downtime_sec is not None for r in recs) else None
        fuel_norm     = recs[0].fuel_norm

        daily_norm_n = report.daily_norm_sec * n
        daily_h_n    = daily_norm_n / 3600 if daily_norm_n > 0 else 0

        fuel_eff_pct = None
        if fuel_norm > 0 and daily_h_n > 0 and total_fuel is not None and total_fuel >= 0:
            fuel_eff_pct = round(total_fuel / (fuel_norm * daily_h_n) * 100, 1)
        output_pct = round(total_engine / daily_norm_n * 100, 1) if daily_norm_n > 0 else None

        type_eff_pct = None
        if group in ('Бульдозеры', 'Погрузчики') and report.bulldozer_norm_sec > 0:
            type_eff_pct = round(total_idle / (report.bulldozer_norm_sec * n) * 100, 1)
        elif group == 'Экскаваторы' and total_dt is not None and report.excavator_norm_sec > 0:
            type_eff_pct = round(total_dt / (report.excavator_norm_sec * n) * 100, 1)
        elif group == 'Самосвалы':
            dt_total = sum(dt_norms.get((recs[0].name, r.shift, r.date), 0) for r in recs)
            if dt_total > 0:
                type_eff_pct = round(total_no_move / dt_total * 100, 1)

        over_bd_str = ''
        if group in ('Бульдозеры', 'Погрузчики') and report.bulldozer_norm_sec > 0:
            ov = total_idle - report.bulldozer_norm_sec * n
            if ov > 0:
                over_bd_str = f'+{secs_to_hhmmss(ov)}'
        elif group == 'Экскаваторы' and total_dt is not None and report.excavator_norm_sec > 0:
            ov = total_dt - report.excavator_norm_sec * n
            if ov > 0:
                over_bd_str = f'+{secs_to_hhmmss(ov)}'

        anomaly_cnt = sum(1 for r in recs if r.has_anomaly)

        ws.append([
            '', recs[0].name, group, date_display, 'Итого',
            secs_to_hhmmss(total_engine), total_fuel, fuel_norm,
            fuel_eff_pct, output_pct, type_eff_pct,
            TYPE_LABEL.get(group, ''), '', over_bd_str,
            total_mileage, total_ref,
            f'Аномалий: {anomaly_cnt}' if anomaly_cnt else '', '',
        ])
        r = ws.max_row
        ws.row_dimensions[r].outline_level = outline_level
        _style_row(r, False, fuel_eff_pct, output_pct, type_eff_pct, over_bd_str, None, is_total=True)
        return fuel_eff_pct, output_pct, type_eff_pct, anomaly_cnt, total_fuel, total_engine

    # ── Build hierarchy: group → vehicle → date → shifts ──────────────
    by_group = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for rec in all_records:
        by_group[rec.group][rec.name][rec.date].append(rec)

    for group_name in sorted(by_group.keys()):
        vehicles    = by_group[group_name]
        total_count = sum(len(s) for dates in vehicles.values() for s in dates.values())

        # Group header row
        ws.append([''] * NCOLS)
        gr = ws.max_row
        ws.merge_cells(f'A{gr}:{get_column_letter(NCOLS)}{gr}')
        gc = ws.cell(row=gr, column=1)
        gc.value       = f'  {group_name.upper()}  ({total_count} записей)'
        gc.font        = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
        gc.fill        = _fill(DARK_BLUE)
        gc.alignment   = _align('left')
        ws.row_dimensions[gr].height = 20

        # Group-level accumulators for group-total row
        g_fuel_eff_vals = []
        g_output_vals   = []
        g_type_eff_vals = []
        g_fuel_sum      = 0.0
        g_engine_h_sum  = 0.0
        g_anomaly_cnt   = 0

        for vehicle_name in sorted(vehicles.keys()):
            dates = vehicles[vehicle_name]
            for date_key in sorted(dates.keys()):
                shifts = sorted(dates[date_key], key=lambda r: r.shift if r.shift else 99)
                multi  = len(shifts) > 1

                for rec in shifts:
                    fe, op, te, r = write_rec_row(rec, outline_level=1 if multi else 0)
                    if fe is not None:
                        g_fuel_eff_vals.append(fe)
                    if op is not None:
                        g_output_vals.append(op)
                    if te is not None:
                        g_type_eff_vals.append(te)
                    if rec.fuel_actual and rec.fuel_actual > 0:
                        g_fuel_sum += rec.fuel_actual
                    g_engine_h_sum += rec.engine_time_sec / 3600
                    if rec.has_anomaly:
                        g_anomaly_cnt += 1

                if multi:
                    write_daily_total(shifts, outline_level=0)

        # Group total row
        g_avg_fuel = round(sum(g_fuel_eff_vals) / len(g_fuel_eff_vals), 1) if g_fuel_eff_vals else None
        g_avg_out  = round(sum(g_output_vals)    / len(g_output_vals),    1) if g_output_vals    else None
        g_avg_type = round(sum(g_type_eff_vals)  / len(g_type_eff_vals),  1) if g_type_eff_vals  else None

        ws.append([
            '', f'ИТОГО {group_name}', '', '', '',
            f'{round(g_engine_h_sum, 1)} ч', round(g_fuel_sum, 1),
            '', g_avg_fuel, g_avg_out, g_avg_type, '', '', '', '', '',
            f'Аномалий: {g_anomaly_cnt}' if g_anomaly_cnt else '', '',
        ])
        tr = ws.max_row
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=tr, column=col)
            c.font      = Font(bold=True, size=11, name='Calibri', color='FFFFFF')
            c.fill      = _fill(MID_BLUE)
            c.alignment = _align()
            c.border    = _border()
        ws.cell(row=tr, column=2).alignment = _align('left')
        f = _pct_fill(g_avg_fuel, 100, 115)
        if f:
            ws.cell(row=tr, column=9).fill = f
        if g_avg_out is not None:
            ws.cell(row=tr, column=10).fill = _fill(GREEN_BG) if g_avg_out >= 80 else _fill(RED_BG)
        f = _pct_fill(g_avg_type, 100, 130)
        if f:
            ws.cell(row=tr, column=11).fill = f

    # ── Column widths ──────────────────────────────────────────────────
    col_widths = [5, 30, 14, 12, 8, 16, 13, 13, 13, 13, 14, 16, 15, 15, 12, 12, 45, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return wb


# ═══════════════════════════════════════════════════════════════════════════════
# ─── Мониторинг Омникомм ──────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

from .models import (
    BreakdownType, FailureCause, MonitoringVehicle, MonitoringRecord,
    MONITORING_GROUPS,
)

MONITORING_GROUP_LIST = [g[0] for g in MONITORING_GROUPS]

GROUP_ICONS = {
    'Самосвалы': 'bi-truck',
    'Экскаваторы': 'bi-tools',
    'Бульдозеры': 'bi-joystick',
    'Погрузчики': 'bi-box-seam',
    'АТЗ': 'bi-fuel-pump',
    'Вспомогательная техника': 'bi-wrench-adjustable',
}


@role_required(GROUP_MONITOR)
def monitoring_index(request):
    today = datetime.date.today()
    user = request.user
    user_section = None if user.is_staff else _get_user_section(user)
    no_section_warning = not user.is_staff and user_section is None

    group_stats = []
    unfilled_groups = []
    for group_name in MONITORING_GROUP_LIST:
        base_qs = MonitoringVehicle.objects.filter(group=group_name, is_active=True)
        if not user.is_staff:
            base_qs = base_qs.filter(section=user_section) if user_section else base_qs.none()
        total = base_qs.count()
        filled_today = MonitoringRecord.objects.filter(
            vehicle__in=base_qs, date=today
        ).count()
        if not user.is_staff and total == 0:
            continue  # Skip groups with no vehicles in this section
        group_stats.append({
            'name': group_name,
            'icon': GROUP_ICONS.get(group_name, 'bi-gear'),
            'total': total,
            'filled_today': filled_today,
        })
        if total > 0 and filled_today < total:
            unfilled_groups.append({
                'name': group_name,
                'filled_today': filled_today,
                'total': total,
                'partial': filled_today > 0,
            })
    return render(request, 'analysis/monitoring/index.html', {
        'group_stats': group_stats,
        'unfilled_groups': unfilled_groups,
        'today': today,
        'no_section_warning': no_section_warning,
        'user_section': user_section,
    })


@role_required(GROUP_MONITOR)
def monitoring_group(request, group):
    if group not in MONITORING_GROUP_LIST:
        return redirect('monitoring_index')

    today = datetime.date.today()

    user = request.user
    user_section = None if user.is_staff else _get_user_section(user)
    no_section_warning = not user.is_staff and user_section is None

    vehicles_qs = MonitoringVehicle.objects.filter(group=group, is_active=True)
    if not user.is_staff:
        vehicles_qs = vehicles_qs.filter(section=user_section) if user_section else vehicles_qs.none()
    vehicles = vehicles_qs.select_related('section').order_by('order', 'name')

    # Dates that already have records (for navigation) — scoped to section
    records_base_qs = MonitoringRecord.objects.filter(vehicle__group=group)
    if not user.is_staff:
        if user_section is not None:
            records_base_qs = records_base_qs.filter(vehicle__section=user_section)
        else:
            records_base_qs = records_base_qs.none()

    recorded_dates = list(
        records_base_qs
        .values_list('date', flat=True)
        .distinct()
        .order_by('-date')[:30]
    )

    # Allowed dates: today always + existing record dates
    allowed_dates = set(recorded_dates) | {today}

    # Determine which date to show — only today or an existing record date
    date_str = request.GET.get('date', '')
    selected_date = today
    if date_str:
        try:
            candidate = datetime.date.fromisoformat(date_str)
            if candidate in allowed_dates:
                selected_date = candidate
            # else silently fall back to today
        except ValueError:
            pass

    # Load existing records for selected date — scoped to section
    existing = {
        r.vehicle_id: r
        for r in records_base_qs.filter(date=selected_date)
        .select_related('breakdown_type', 'failure_cause')
    }

    breakdown_types = BreakdownType.objects.all()
    failure_causes = FailureCause.objects.all()

    is_excavator = group == 'Экскаваторы'
    is_atz = group == 'АТЗ'

    first_record = next(iter(existing.values()), None)
    prefill_author = (first_record.author if first_record else '') or request.user.username

    context = {
        'group': group,
        'group_icon': GROUP_ICONS.get(group, 'bi-gear'),
        'no_section_warning': no_section_warning,
        'vehicles': vehicles,
        'existing': existing,
        'selected_date': selected_date,
        'today': today,
        'recorded_dates': recorded_dates,
        'breakdown_types': breakdown_types,
        'failure_causes': failure_causes,
        'is_excavator': is_excavator,
        'is_atz': is_atz,
        'prefill_author': prefill_author,
    }
    return render(request, 'analysis/monitoring/group.html', context)


@role_required(GROUP_MONITOR)
@require_POST
def monitoring_save(request, group):
    if group not in MONITORING_GROUP_LIST:
        return redirect('monitoring_index')

    today = datetime.date.today()
    date_str = request.POST.get('date', '')
    try:
        record_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        messages.error(request, 'Неверная дата.')
        return redirect('monitoring_group', group=group)

    # Only allow saving for today or dates that already have records
    existing_dates = set(
        MonitoringRecord.objects.filter(vehicle__group=group)
        .values_list('date', flat=True)
        .distinct()
    )
    if record_date != today and record_date not in existing_dates:
        messages.error(request, 'Нельзя создавать записи задним числом.')
        return redirect('monitoring_group', group=group)

    user = request.user
    user_section = None if user.is_staff else _get_user_section(user)
    if not user.is_staff and user_section is None:
        messages.error(request, 'Вам не назначен участок. Обратитесь к администратору.')
        return redirect('monitoring_group', group=group)

    author = (request.POST.get('author', '').strip()) or request.user.username
    is_excavator = group == 'Экскаваторы'
    is_atz = group == 'АТЗ'

    vehicles = MonitoringVehicle.objects.filter(group=group, is_active=True)
    if not user.is_staff and user_section is not None:
        vehicles = vehicles.filter(section=user_section)
    saved_count = 0

    with transaction.atomic():
        for v in vehicles:
            # Defense-in-depth: skip any vehicle not belonging to the user's section
            if not user.is_staff and user_section is not None and v.section != user_section:
                continue

            prefix = f'v_{v.pk}_'

            def chk(field):
                return request.POST.get(prefix + field) == '1'

            sensor_rpm = chk('rpm')
            sensor_dut = chk('dut')
            sensor_gps = chk('gps')
            sensor_gsm = chk('gsm')
            sensor_arrow = chk('arrow') if is_excavator else None
            sensor_cube_port = chk('cube_port') if is_atz else None
            sensor_uss = chk('uss') if is_atz else None

            bd_id = request.POST.get(prefix + 'breakdown_type') or None
            fc_id = request.POST.get(prefix + 'failure_cause') or None
            note = request.POST.get(prefix + 'note', '').strip()

            # Check if any sensor is faulty
            sensors = [sensor_rpm, sensor_dut, sensor_gps, sensor_gsm]
            if is_excavator:
                sensors.append(sensor_arrow)
            if is_atz:
                sensors += [sensor_cube_port, sensor_uss]
            any_fault = any(s is False for s in sensors)
            if not any_fault:
                bd_id = None
                fc_id = None

            MonitoringRecord.objects.update_or_create(
                vehicle=v,
                date=record_date,
                defaults={
                    'author': author,
                    'sensor_rpm': sensor_rpm,
                    'sensor_dut': sensor_dut,
                    'sensor_gps': sensor_gps,
                    'sensor_gsm': sensor_gsm,
                    'sensor_arrow': sensor_arrow,
                    'sensor_cube_port': sensor_cube_port,
                    'sensor_uss': sensor_uss,
                    'breakdown_type_id': int(bd_id) if bd_id else None,
                    'failure_cause_id': int(fc_id) if fc_id else None,
                    'note': note,
                }
            )
            saved_count += 1

    messages.success(request, f'Сохранено записей: {saved_count} за {record_date.strftime("%d.%m.%Y")}.')
    return redirect(f'{request.path_info.replace("save/", "")}?date={record_date.isoformat()}')


@role_required(GROUP_MONITOR)
def monitoring_analytics(request):
    import json as _json
    from collections import Counter, defaultdict

    metric = request.GET.get('metric', 'breakdown')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    group_filter = request.GET.get('group', '')

    # Checkbox filters
    selected_breakdown_ids = [
        int(x) for x in request.GET.getlist('breakdown_ids') if x.isdigit()
    ]
    selected_failure_ids = [
        int(x) for x in request.GET.getlist('failure_ids') if x.isdigit()
    ]

    user = request.user
    user_section = None if user.is_staff else _get_user_section(user)

    qs = MonitoringRecord.objects.select_related(
        'vehicle', 'vehicle__section', 'breakdown_type', 'failure_cause'
    )
    if not user.is_staff:
        if user_section is not None:
            qs = qs.filter(vehicle__section=user_section)
        else:
            qs = qs.none()

    if date_from:
        try:
            qs = qs.filter(date__gte=datetime.date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(date__lte=datetime.date.fromisoformat(date_to))
        except ValueError:
            pass
    if group_filter:
        qs = qs.filter(vehicle__group=group_filter)

    # Evaluate once — reuse for group/vehicle summaries
    all_records = list(qs)

    # All fault records (no checkbox filter) — for group/vehicle table
    all_fault_records = [r for r in all_records if r.has_fault()]

    # Fault records filtered by checkbox selection — for chart/metric table
    chart_fault_records = all_fault_records
    if metric == 'breakdown' and selected_breakdown_ids:
        chart_fault_records = [
            r for r in all_fault_records
            if r.breakdown_type_id in selected_breakdown_ids
        ]
    elif metric == 'failure' and selected_failure_ids:
        chart_fault_records = [
            r for r in all_fault_records
            if r.failure_cause_id in selected_failure_ids
        ]

    if metric == 'breakdown':
        id_counts = Counter()
        id_vehicles = defaultdict(set)
        id_names = {}
        for r in chart_fault_records:
            pk = r.breakdown_type_id
            name = r.breakdown_type.name if r.breakdown_type else '(не указан)'
            id_counts[pk] += 1
            id_vehicles[pk].add(r.vehicle_id)
            id_names[pk] = name
    else:
        id_counts = Counter()
        id_vehicles = defaultdict(set)
        id_names = {}
        for r in chart_fault_records:
            pk = r.failure_cause_id
            name = r.failure_cause.name if r.failure_cause else '(не указана)'
            id_counts[pk] += 1
            id_vehicles[pk].add(r.vehicle_id)
            id_names[pk] = name

    chart_data_raw = sorted(id_counts.items(), key=lambda x: -x[1])
    chart_table = [
        {
            'id': pk if pk is not None else 'none',
            'name': id_names[pk],
            'count': cnt,
            'vehicle_count': len(id_vehicles[pk]),
        }
        for pk, cnt in chart_data_raw
    ]
    chart_labels = [id_names[pk] for pk, cnt in chart_data_raw]
    chart_values = [cnt for pk, cnt in chart_data_raw]
    chart_ids = [pk if pk is not None else 'none' for pk, cnt in chart_data_raw]

    # Group + per-vehicle summary
    group_summary = defaultdict(lambda: {'total': 0, 'faults': 0})
    vehicle_summary = defaultdict(lambda: {'total': 0, 'faults': 0})

    for r in all_records:
        grp = r.vehicle.group
        key = (grp, r.vehicle.pk, r.vehicle.name)
        group_summary[grp]['total'] += 1
        vehicle_summary[key]['total'] += 1
        if r.has_fault():
            group_summary[grp]['faults'] += 1
            vehicle_summary[key]['faults'] += 1

    group_table = []
    for grp, gdata in sorted(group_summary.items()):
        gt = gdata['total']
        gf = gdata['faults']
        vehicle_rows = []
        for (g, vpk, vname), vdata in sorted(vehicle_summary.items(), key=lambda x: x[0][2]):
            if g != grp:
                continue
            vt = vdata['total']
            vf = vdata['faults']
            vehicle_rows.append({
                'name': vname,
                'total': vt,
                'faults': vf,
                'ok': vt - vf,
                'fault_pct': round(vf / vt * 100, 1) if vt else 0,
            })
        group_table.append({
            'group': grp,
            'total': gt,
            'faults': gf,
            'ok': gt - gf,
            'fault_pct': round(gf / gt * 100, 1) if gt else 0,
            'vehicles': vehicle_rows,
        })

    all_breakdown_types = list(BreakdownType.objects.all())
    all_failure_causes = list(FailureCause.objects.all())

    context = {
        'metric': metric,
        'date_from': date_from,
        'date_to': date_to,
        'group_filter': group_filter,
        'group_list': MONITORING_GROUP_LIST,
        'selected_breakdown_ids': selected_breakdown_ids,
        'selected_failure_ids': selected_failure_ids,
        'all_breakdown_types': all_breakdown_types,
        'all_failure_causes': all_failure_causes,
        'chart_labels_json': _json.dumps(chart_labels, ensure_ascii=False),
        'chart_values_json': _json.dumps(chart_values),
        'chart_ids_json': _json.dumps(chart_ids),
        'chart_table': chart_table,
        'group_table': group_table,
        'total_records': len(all_records),
        'total_faults': len(all_fault_records),
    }
    return render(request, 'analysis/monitoring/analytics.html', context)


# ─── Drill-down: vehicles for a specific fault type ───────────────────────────

@role_required(GROUP_MONITOR, GROUP_ANALYST)
def monitoring_fault_drill(request):
    metric = request.GET.get('metric', 'breakdown')
    item_id = request.GET.get('item_id', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    group_filter = request.GET.get('group', '')

    user = request.user
    user_section = None if user.is_staff else _get_user_section(user)

    qs = MonitoringRecord.objects.select_related(
        'vehicle', 'vehicle__section', 'breakdown_type', 'failure_cause'
    )
    if not user.is_staff:
        if user_section is not None:
            qs = qs.filter(vehicle__section=user_section)
        else:
            qs = qs.none()

    if date_from:
        try:
            qs = qs.filter(date__gte=datetime.date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(date__lte=datetime.date.fromisoformat(date_to))
        except ValueError:
            pass
    if group_filter:
        qs = qs.filter(vehicle__group=group_filter)

    item_name = ''
    if metric == 'breakdown':
        if item_id == 'none' or not item_id:
            qs = qs.filter(breakdown_type__isnull=True)
            item_name = '(не указан)'
        else:
            try:
                bt = BreakdownType.objects.get(pk=int(item_id))
                qs = qs.filter(breakdown_type=bt)
                item_name = bt.name
            except (BreakdownType.DoesNotExist, ValueError):
                qs = qs.none()
                item_name = '—'
    else:
        if item_id == 'none' or not item_id:
            qs = qs.filter(failure_cause__isnull=True)
            item_name = '(не указана)'
        else:
            try:
                fc = FailureCause.objects.get(pk=int(item_id))
                qs = qs.filter(failure_cause=fc)
                item_name = fc.name
            except (FailureCause.DoesNotExist, ValueError):
                qs = qs.none()
                item_name = '—'

    records = [r for r in qs.order_by('-date', 'vehicle__name') if r.has_fault()]

    back_params = [f'metric={metric}']
    if date_from:
        back_params.append(f'date_from={date_from}')
    if date_to:
        back_params.append(f'date_to={date_to}')
    if group_filter:
        back_params.append(f'group={group_filter}')
    back_url = reverse('monitoring_analytics') + '?' + '&'.join(back_params)

    context = {
        'metric': metric,
        'item_id': item_id,
        'item_name': item_name,
        'date_from': date_from,
        'date_to': date_to,
        'group_filter': group_filter,
        'records': records,
        'back_url': back_url,
        'total': len(records),
    }
    return render(request, 'analysis/monitoring/fault_drill.html', context)


# ─── Reference CRUD: MonitoringVehicle ───────────────────────────────────────

@staff_required
def monitoring_vehicles(request):
    vehicles = MonitoringVehicle.objects.select_related('section').order_by('group', 'order', 'name')
    sections = Section.objects.all()
    return render(request, 'analysis/monitoring/vehicles.html', {
        'vehicles': vehicles,
        'sections': sections,
        'group_list': MONITORING_GROUP_LIST,
    })


@staff_required
def monitoring_vehicle_create(request):
    sections = Section.objects.all()
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        group = request.POST.get('group', '')
        section_id = request.POST.get('section') or None
        order = int(request.POST.get('order', 0) or 0)
        is_active = request.POST.get('is_active') == '1'
        if name and group in MONITORING_GROUP_LIST:
            MonitoringVehicle.objects.create(
                name=name, group=group,
                section_id=section_id, order=order, is_active=is_active,
            )
            messages.success(request, f'Техника «{name}» добавлена.')
            return redirect('monitoring_vehicles')
        messages.error(request, 'Заполните название и выберите группу.')
    return render(request, 'analysis/monitoring/vehicle_form.html', {
        'sections': sections,
        'group_list': MONITORING_GROUP_LIST,
        'obj': None,
    })


@staff_required
def monitoring_vehicle_edit(request, pk):
    v = get_object_or_404(MonitoringVehicle, pk=pk)
    sections = Section.objects.all()
    if request.method == 'POST':
        v.name = request.POST.get('name', '').strip()
        v.group = request.POST.get('group', v.group)
        v.section_id = request.POST.get('section') or None
        v.order = int(request.POST.get('order', 0) or 0)
        v.is_active = request.POST.get('is_active') == '1'
        v.save()
        messages.success(request, f'Техника «{v.name}» обновлена.')
        return redirect('monitoring_vehicles')
    return render(request, 'analysis/monitoring/vehicle_form.html', {
        'sections': sections,
        'group_list': MONITORING_GROUP_LIST,
        'obj': v,
    })


@staff_required
def monitoring_vehicle_delete(request, pk):
    v = get_object_or_404(MonitoringVehicle, pk=pk)
    if request.method == 'POST':
        name = v.name
        v.delete()
        messages.success(request, f'Техника «{name}» удалена.')
        return redirect('monitoring_vehicles')
    return render(request, 'analysis/monitoring/vehicle_confirm_delete.html', {'obj': v})


# ─── Reference CRUD: BreakdownType ───────────────────────────────────────────

@staff_required
def monitoring_breakdowns(request):
    items = BreakdownType.objects.all()
    return render(request, 'analysis/monitoring/breakdowns.html', {'items': items})


@staff_required
def monitoring_breakdown_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            BreakdownType.objects.create(name=name)
            messages.success(request, f'Вид поломки «{name}» добавлен.')
            return redirect('monitoring_breakdowns')
        messages.error(request, 'Введите название.')
    return render(request, 'analysis/monitoring/breakdown_form.html', {'obj': None})


@staff_required
def monitoring_breakdown_edit(request, pk):
    obj = get_object_or_404(BreakdownType, pk=pk)
    if request.method == 'POST':
        obj.name = request.POST.get('name', '').strip()
        obj.save()
        messages.success(request, f'Обновлено: «{obj.name}».')
        return redirect('monitoring_breakdowns')
    return render(request, 'analysis/monitoring/breakdown_form.html', {'obj': obj})


@staff_required
def monitoring_breakdown_delete(request, pk):
    obj = get_object_or_404(BreakdownType, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Вид поломки удалён.')
        return redirect('monitoring_breakdowns')
    return render(request, 'analysis/monitoring/breakdown_confirm_delete.html', {'obj': obj})


# ─── Reference CRUD: FailureCause ────────────────────────────────────────────

@staff_required
def monitoring_failures(request):
    items = FailureCause.objects.all()
    return render(request, 'analysis/monitoring/failures.html', {'items': items})


@staff_required
def monitoring_failure_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            FailureCause.objects.create(name=name)
            messages.success(request, f'Причина «{name}» добавлена.')
            return redirect('monitoring_failures')
        messages.error(request, 'Введите название.')
    return render(request, 'analysis/monitoring/failure_form.html', {'obj': None})


@staff_required
def monitoring_failure_edit(request, pk):
    obj = get_object_or_404(FailureCause, pk=pk)
    if request.method == 'POST':
        obj.name = request.POST.get('name', '').strip()
        obj.save()
        messages.success(request, f'Обновлено: «{obj.name}».')
        return redirect('monitoring_failures')
    return render(request, 'analysis/monitoring/failure_form.html', {'obj': obj})


@staff_required
def monitoring_failure_delete(request, pk):
    obj = get_object_or_404(FailureCause, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Причина неисправности удалена.')
        return redirect('monitoring_failures')
    return render(request, 'analysis/monitoring/failure_confirm_delete.html', {'obj': obj})
